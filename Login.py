import flet as ft
from flet import (
    AppBar,
    ElevatedButton,
    Page,
    Text,
    View,
    colors,
    Column,
    Container,
    LinearGradient,
    alignment,
    border_radius,
    padding,
    Image,
    UserControl,
    Row,
    IconButton,
    margin,
    Card,
    TextField,
    FilledButton,
    SnackBar
)
import time
import threading
import sys
import pyodbc
import socket
import base64
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import pikepdf
import win32com.client as win32
import os
from datetime import datetime, timedelta,date
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
import logging
global nombre
import re
# Configuración del registro
logging.basicConfig(
    filename="error_log.log",  # Nombre del archivo de log
    level=logging.ERROR,       # Nivel de registro (ERROR, WARNING, INFO, DEBUG)
    format="%(asctime)s - %(levelname)s - %(message)s"  # Formato del mensaje
)
class Variables:
    def __init__(self):
        self.KEY_ENCRYPT_DECRYPT = "r3c6rs0sm4t3r14l3sj6m4p4mm4z4tl4ns1n4l04l4p13ld3lm4rr3c6rs0sm4t3r14l3sj6m4p4mm4z4tl4ns1n4l04l4p13ld3lm4r"
        self.periodos=""
        self.semana=["LUNES","MARTES","MIERCOLES","JUEVES","VIERNES","SABADO","DOMINGO"]
        self.dias=[]
        self.asis={}
        self.role=""
        self.multireg=False
        self.fechas_seleccionadas = {}
        self.fechas=[]
        self.year=datetime.now().year
        self.month=datetime.now().month
        self.day=datetime.now().day
        self.todos={}
        self.dias_festivos=[date(self.year,1,1),date(self.year,5,1),date(self.year,9,16),date(self.year,12,25)]
        self.dias_semana={0:"LUNES",1:"MARTES",2:"MIERCOLES",3:"JUEVES",4:"VIERNES",5:"SABADO",6:"DOMINGO",}
        self.excel_file =  os.path.dirname(__file__)+"\\Formatollenado.xlsx"
        self.pdf_file =  os.path.dirname(sys.executable)+"\\output.pdf"
        self.HE_entries={}
        self.DT_entries={}
        self.TE_entries={}
        self.nombreusuario=""
    def obtener_todos(self):
        return self.periodos,self.dias,self.asis,self.multireg,self.todos,self.excel_file,self.pdf_file,self.HE_entries,self.DT_entries,self.TE_entries
    def obtener_llave(self):
        return self.KEY_ENCRYPT_DECRYPT
    def agregar_festivo(self, year, month,day):
        self.dias_festivos.append(date(year,month,day))
    def agregar_usuario(self, user):
        self.nombreusuario=user
    def agregar_fecha(self,fecha,dt,ds,f):
        if f:
            self.fechas_seleccionadas[dt,ds]=fecha
        else:
            del self.fechas_seleccionadas[dt,ds]
    def obtener_usuario(self):
        return self.nombreusuario
    def obtener_fechas(self):
        return self.year,self.month,self.day,self.dias_festivos,self.fechas_seleccionadas,self.fechas
    def obtener_semana(self):
        return self.dias_semana,self.semana

buttonClicked  = False # Bfore first click
def verificar_Permisos(usuario):
    try:
        db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
        db.conectar()
        id=db.ejecutar_consulta("SELECT ID_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE NOMBRE_USUARIO = ?",usuario)
        cursor=db.cursor
        try:
            cursor.execute("{CALL spAccesoSistemasPermisos (?)}",id[0])
        except Exception as e:
            logging.error(f"No se pudo Verificar: {e}")
            raise
        # Obtener los resultados
        rows = cursor.fetchall()
        db.commit()
        db.cerrar()
        i=0
        for row in rows:
            if "VALIDAR" in str(row):
                if "True" in str(row[11]):
                    return False
            if "ADMIN" in str(row):
                if "True" in str(row[11]):
                    return "False"
        return True
    except Exception as e:
        logging.error("Error", f"No se pudo Verificar: {e}")
def verificar_Permisos2(usuario):
    try:
        db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
        db.conectar()
        id=db.ejecutar_consulta("SELECT ID_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE NOMBRE_USUARIO = ?",usuario)
        cursor=db.cursor
        try:
            cursor.execute("{CALL spAccesoSistemasPermisos (?)}",id[0])
        except Exception as e:
            logging.error(f"No se pudo Verificar: {e}")
            raise
        # Obtener los resultados
        rows = cursor.fetchall()
        db.commit()
        db.cerrar()
        i=0
        for row in rows:
            print(row)
            if "ADMIN" in str(row):
                if "True" in str(row[11]):
                    return True
        return False
    except Exception as e:
        logging.error("Error", f"No se pudo Verificar: {e}")
def calcular_fechas():
    var=Variables()
    year,month,day,dias_festivos,_,__=var.obtener_fechas()
    if ".333" in str(year/6):
        if date(year, 10, 1) not in var.dias_festivos:
            var.agregar_festivo(year,10,1)
    d = date(year, 2, 1)
    offset = 0-d.weekday() #weekday = 0 means monday
    if offset < 0:
        offset+=7
    if date(year,2,1+offset) not in var.dias_festivos:
        var.agregar_festivo(year,2,1+offset)
    d=date(year,3,14)
    offset=0+-d.weekday()
    if offset < 0:
        offset+=7
    if date(year,3,14+offset) not in var.dias_festivos:
        var.agregar_festivo(year,3,14+offset)
    d=date(year,11,14)
    offset=0+-d.weekday()
    if offset < 0:
        offset+=7
    if date(year,11,14+offset) not in var.dias_festivos:
        var.agregar_festivo(year,11,14+offset)
    return dias_festivos
def checar_aprovacion(codigoEmpleado,periodo):
    try:
        # Conectar a la base de datos
        db = ConexionBD(host="148.200.128.13", database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
        db.conectar()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        resultado = db.ejecutar_consulta("SELECT Aprobacion FROM Prenomina WHERE Codigo_Empleado = ? And Periodo= ?",(codigoEmpleado,periodo))
        if "0" in str(resultado) or resultado == None:
            return False
        db.cerrar()
        return True
    except Exception as e:
        logging.error("Error", f"No se pudo Verificar: {e}")
def check_dias(dia,codigoempleado,periodo):
    try:
        db = ConexionBD(host="148.200.128.13", database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
        db.conectar()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        resultado = db.ejecutar_consulta("SELECT Dia_Asistencia FROM Prenomina Where Codigo_Empleado=? AND Dia_Semana=? AND Periodo=?",(codigoempleado,dia,periodo))
        x=len(resultado)
        print(resultado,x)
        if x==0:
            if dia == "SABADO" or dia=="DOMINGO":
                return False
            return True
        if "None" in str(resultado):
            return False
        elif "/" in str(resultado):
            return True
        db.cerrar()
    except Exception as e:
        logging.error("Error", f"No se pudo Verificar: {e}")
def obtener_ASIGNED_departamentos(id):
    db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
    db.conectar()
    resultados=db.ejecutar_consulta("SELECT CLAVE_DEPARTAMENTO FROM HIS_SISTEMAS_DEPUSER Where ID_USUARIO=?",(id))
    depars=[]
    for i in resultados:
        depars.append(str(i[0]))
    db.cerrar()
    return depars
def obtener_ALL_departamentos():
    db = ConexionBD(host="148.200.128.13",database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
    db.conectar()
    resultados=db.ejecutar_consulta("SELECT CLAVE_DEPARTAMENTO FROM DEPARTAM")
    depars=[]
    for i in resultados:
        depars.append(str(i[0]))
    db.cerrar()
    return depars
def GET_USER():
    try:
        db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
        db.conectar()
        resultado=db.ejecutar_consulta("SELECT ID_USUARIO FROM HIS_SISTEMAS_PERMISOS WHERE ID_SISTEMA = 12")
        id=resultado
        iD={}
        usuarios=[]
        for ids in id:
            resultado2=db.ejecutar_consulta("SELECT NOMBRE_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE ID_USUARIO = ?",ids[0])
            usuarios.append(str(resultado2[0][0]))
            iD[str(resultado2[0][0])]=(str(ids[0]))
        db.cerrar()
        return iD,usuarios
    except Exception as e:
        logging.error("Error", f"Error al ejecutar el Proceso alamcenado: {e}")
def UPDATE_USER(id,clave):
    try:
        db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
        db.conectar()
        if not id or not clave:
            raise ValueError("ID o clave no válidos.")
        for clavedep in clave:
            resultado=db.ejecutar_consulta("SELECT ID_USUARIO FROM HIS_SISTEMAS_DEPUSER WHERE ID_USUARIO = ? AND CLAVE_DEPARTAMENTO=?",(id,str(clavedep)))
            if str(id) in resultado:
                print("HI")
            else:
                db.ejecutar_consulta("INSERT INTO HIS_SISTEMAS_DEPUSER VALUES(?,?)",(id,str(clavedep)),commit=True)
        resultado2=db.ejecutar_consulta("SELECT NOMBRE_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE ID_USUARIO = ?",id)
        db.cerrar()
        return resultado2
    except Exception as e:
        logging.error("Error", f"Error al ejecutar el Proceso alamcenado: {e}")
def DEL_USER(id,clave):
    try:
        db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
        db.conectar()
        for clavedep in clave:
            db.ejecutar_consulta("DELETE FROM HIS_SISTEMAS_DEPUSER WHERE ID_USUARIO = ? AND CLAVE_DEPARTAMENTO=?",(id,str(clavedep)),commit=True)
        resultado2=db.ejecutar_consulta("SELECT NOMBRE_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE ID_USUARIO = ?",id)
        db.cerrar()
        return resultado2
    except Exception as e:
        logging.error("Error", f"Error al ejecutar el Proceso alamcenado: {e}")
def callback():
    global buttonClicked
    buttonClicked = not buttonClicked 
def obtener_quincenas():
    var=Variables()
    year,month,day,dias_festivos,_,__=var.obtener_fechas()
    a=[]
    if day <= 15:
        # Primera quincena (del 1 al 15)
        inicio = date(year, month, 1)
        fin = date(year, month, 15)
        for i in range(15):
            a.append((inicio+timedelta(i)).weekday())
    else:
        # Segunda quincena (del 16 al último día del mes)
        inicio = date(year, month, 16)
        ultimo_dia = (inicio.replace(month=month % 12 + 1, day=1) - timedelta(days=1)).day
        fin = date(year, month, ultimo_dia)
        for i in range(ultimo_dia-15):
            a.append((inicio+timedelta(i)).weekday())    
    return a
def sortdias(dias):
    var=Variables()
    dias_semana,semana=var.obtener_semana()
    a={}
    for i,dia in enumerate(dias):
        if dia ==5 or dia ==6:
            a.update({(dias_semana[dia],i):False})
        else:
            a.update({(dias_semana[dia],i):True})
    return a
def obtener_empleados(depar,tipo):
    if tipo== "Confianza":
        tipoc=2
    else:
        tipoc=1
    db = ConexionBD(host="148.200.128.13",database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
    db.conectar()
    resultados=db.ejecutar_consulta("SELECT TRABAJAD.CLAVE_TRABAJADOR,CLAVE_TIPO_NOMINA, NOMBRE, PATERNO, MATERNO, DESCANSO1, DESCANSO2,CLAVE_DEPARTAMENTO FROM TRABAJAD INNER JOIN TRAHISDE ON TRAHISDE.CLAVE_TRABAJADOR=TRABAJAD.CLAVE_TRABAJADOR WHERE FECHA_F='2100-12-31' AND CLAVE_DEPARTAMENTO=? AND CLAVE_TIPO_NOMINA=?",(depar,tipoc))
    db.cerrar()
    return resultados
def obtener_empleados_search(depar,tipo,ID):
    if tipo== "Confianza":
        tipoc=2
    else:
        tipoc=1
    db = ConexionBD(host="148.200.128.13",database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
    db.conectar()
    resultados=db.ejecutar_consulta("SELECT TRABAJAD.CLAVE_TRABAJADOR,CLAVE_TIPO_NOMINA, NOMBRE, PATERNO, MATERNO, DESCANSO1, DESCANSO2,CLAVE_DEPARTAMENTO FROM TRABAJAD INNER JOIN TRAHISDE ON TRAHISDE.CLAVE_TRABAJADOR=TRABAJAD.CLAVE_TRABAJADOR WHERE FECHA_F='2100-12-31' AND CLAVE_DEPARTAMENTO=? AND CLAVE_TIPO_NOMINA=? AND TRABAJAD.CLAVE_TRABAJADOR LIKE '%'+?",(depar,tipoc,ID))
    db.cerrar()
    if ID=="":
        obtener_empleados(depar,tipo)
    return resultados
def obtener_periodo(x):
    hoy=date.today()
    print(hoy)
    if x==1:
        db = ConexionBD(host="148.200.128.13",database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
        db.conectar()
        periodo =db.ejecutar_consulta("SELECT CLAVE_PERIODO, CLAVE_TIPO_NOMINA, DESCRIPCION FROM PERIODO WHERE CLAVE_TIPO_NOMINA=1 AND ? BETWEEN FECHA_I AND FECHA_F",str(hoy))
        db.cerrar()
        return periodo
    else:
        db = ConexionBD(host="148.200.128.13",database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
        db.conectar()
        periodo =db.ejecutar_consulta("SELECT CLAVE_PERIODO, CLAVE_TIPO_NOMINA, DESCRIPCION FROM PERIODO WHERE CLAVE_TIPO_NOMINA=2 AND ? BETWEEN FECHA_I AND FECHA_F",str(hoy))
        db.cerrar()
        return periodo
def obtener_periodos_mes(x,hoy):
    print(x)
    if x=="Sindicato":
        db = ConexionBD(host="148.200.128.13",database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
        db.conectar()
        periodo =db.ejecutar_consulta("SELECT CLAVE_PERIODO, CLAVE_TIPO_NOMINA, DESCRIPCION,FECHA_I FROM PERIODO WHERE CLAVE_TIPO_NOMINA=1 AND DATEPART(MM,FECHA_I) = ?",hoy)
        db.cerrar()
        return periodo
    else:
        db = ConexionBD(host="148.200.128.13",database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
        db.conectar()
        periodo =db.ejecutar_consulta("SELECT CLAVE_PERIODO, CLAVE_TIPO_NOMINA, DESCRIPCION,FECHA_I FROM PERIODO WHERE CLAVE_TIPO_NOMINA=2 AND DATEPART(MM,FECHA_I) = ?",hoy)
        db.cerrar()
        return periodo
def excel():
    var=Variables()
    year,month,day,dias_festivos,fechas_seleccionadas,fechas=var.obtener_fechas()
    hoy = datetime.now()
    fechas = []
    # Calcular la fecha del día seleccionado
    for i in range(7):
        fecha = hoy - timedelta(days=hoy.weekday()) + timedelta(days=i)
        fecha_formateada = fecha.strftime('%d/%m/%Y')
        fechas.append(fecha_formateada)
    # Crear un nuevo libro de Excel o cargar uno existente
    ruta_excel = os.path.dirname(__file__)+"\\Formato.xlsx"
    # Verificar si el archivo existe y cargarlo, de lo contrario, crear uno nuevo
    try:
        workbook = load_workbook(ruta_excel)
    except FileNotFoundError:
        logging.warning("El archivo de Excel no existe, se creará uno nuevo.")
        workbook = Workbook()
    # Obtener la hoja de trabajo (el índice empieza en cero, o usa el nombre de la hoja)
    sheet = workbook.active  # O usa workbook["NombreHoja"] para acceder a una hoja específica
    # Asignar fechas a las celdas correspondientes
    for i in range(7):
        column_letter = get_column_letter(5 + i)  # E es la columna 5
        sheet[f"{column_letter}3"] = fechas[i]
    # Guardar el archivo de Excel actualizado
    try:
        workbook.save(ruta_excel)
    except PermissionError as e:
        logging.error(f"No se pudo guardar el archivo de Excel: {e}")
    # Configuración de la página
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.printGridlines = True
    workbook.close()
    hoy=datetime.now().day
    hoy2=datetime.now()
    fechasQ = []
    dias=obtener_quincenas()
    # Calcular la fecha del día seleccionado
    for i,fec in enumerate(dias):
        if day <=15:
            fecha = hoy2 - timedelta(days=day) + timedelta(days=i+1)
        elif day>15:
            fecha=datetime(year,month,16)+timedelta(days=i)
        # Formatear la fecha como DD/MM/AAAA
        fecha_formateada = fecha.strftime('%d/%m/%Y')
        fechasQ.append(fecha_formateada)
    # Crear un nuevo libro de Excel o cargar uno existente
    ruta_excel =os.path.dirname(__file__)+"\\Formato - copia.xlsx"
    # Verificar si el archivo existe y cargarlo, de lo contrario, crear uno nuevo
    try:
        workbook = load_workbook(ruta_excel)
    except FileNotFoundError:
        workbook = Workbook()
    # Obtener la hoja de trabajo (el índice empieza en cero, o usa el nombre de la hoja)
    sheet = workbook.active  # O usa workbook["NombreHoja"] para acceder a una hoja específica
    # Asignar fechas a las celdas correspondientes
    for i,fec in enumerate(dias):
        column_letter = get_column_letter(5 + i)  # E es la columna 5
        sheet[f"{column_letter}3"] = fechasQ[i]
    # Guardar el archivo de Excel actualizado
    workbook.save(ruta_excel)
    # Configuración de la página
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.printGridlines = True
    workbook.close()
excel()
def verificar_asistencias(codigoEmpleado,periodo):
    try:
        # Conectar a la base de datos
        db = ConexionBD(host="148.200.128.13", database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
        db.conectar()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        resultado = db.ejecutar_consulta("SELECT Codigo_Empleado,Dia_Asistencia,Horas_Extra,Turnos_Extras,Descansos_Trabajados FROM Prenomina WHERE Codigo_Empleado=? AND Periodo=?",(codigoEmpleado,periodo))
        asistencia_modificada = []
        for fila in resultado:
            print(fila[1])
            if "/" in str(fila[1]): 
                dia_asistencia = 1
            else:
                dia_asistencia = 0
            asistencia_modificada.append((fila[0], dia_asistencia, fila[2], fila[3], fila[4]))
        db.cerrar()
        return asistencia_modificada if asistencia_modificada else None  # Asegurarse de devolver None si no hay datos
    except Exception as e:
        logging.error("Error", f"Verificando: {e}")
def excel_add(id,depar,tipo,periodo):
    # Cargar el archivo de Excel
    ruta= os.path.dirname(__file__)
    if not os.path.exists(os.path.dirname(sys.executable)+"\\Reportes"):
        os.mkdir(os.path.dirname(sys.executable)+"\\Reportes")
    if tipo=="Sindicato":
        ruta_excel =  os.path.dirname(__file__)+"\\Formato.xlsx"
    else:
        ruta_excel =  os.path.dirname(__file__)+"\\Formato - copia.xlsx"
    try:
        workbook = load_workbook(ruta_excel)
    except FileNotFoundError:
        logging.warning(f"El archivo {ruta_excel} no existe. Creando uno nuevo.")
        workbook = Workbook()
    # Obtener la hoja de trabajo
    sheet = workbook.active 
    x=0
    empleados = obtener_empleados(depar,tipo)
    nombres = []
    ape=[]
    ape2=[]
    code=[]
    # Extraer el nombre completo del empleado
    for empleado in empleados:
        code.append(empleado[0])
        nombre_completo = f"{empleado[2]}"
        nombres.append(nombre_completo)
        ap=f"{empleado[3]}"
        ape.append(ap)
        ap2=f"{empleado[4]}"
        ape2.append(ap2)
        x+=1
    if len(empleados) == 0:
        logging.warning("No hay empleados disponibles para escribir en Excel.")
        return False
    if tipo=="Sindicato":
        for i in range(4, x + 4):
            Dato = verificar_asistencias(code[(i - 4)],periodo)
            if Dato is not None:
                sheet["A" + str(i)] = str(Dato[0][0])  # Código de empleado
                sheet["B" + str(i)] = str(nombres[i - 4])  # Nombre completo del empleado
                sheet["C" + str(i)] = str(ape[i - 4])  # Nombre completo del empleado
                sheet["D" + str(i)] = str(ape2[i - 4])  # Nombre completo del empleado
                # Solo asignar si hay datos en Dato[0] y Dato[0][1]
                sheet["E" + str(i)] = 1 if Dato[0][1] > 0 else 0  # Dias_Asistencia (0 o 1)
                sheet["L" + str(i)] = str(Dato[0][2])  # Horas_Extra
                sheet["M" + str(i)] = str(Dato[0][3])  # Turnos_Extras
                sheet["N" + str(i)] = str(Dato[0][4])  # Descansos_Trabajados
                # Asegúrate de verificar la longitud de Dato[1] y demás
                sheet["F" + str(i)] = str(Dato[1][1])  # Datos de asistencia
                sheet["O" + str(i)] = str(Dato[1][2])  # Más datos
                sheet["P" + str(i)] = str(Dato[1][3])  # Más datos
                sheet["Q" + str(i)] = str(Dato[1][4])  # Más datos
                sheet["G" + str(i)] = str(Dato[2][1])  # Datos adicionales
                sheet["R" + str(i)] = str(Dato[2][2])  # Más datos
                sheet["S" + str(i)] = str(Dato[2][3])  # Más datos
                sheet["T" + str(i)] = str(Dato[2][4])  # Más datos
                sheet["H" + str(i)] = str(Dato[3][1])  # Más datos
                sheet["U" + str(i)] = str(Dato[3][2])  # Más datos
                sheet["V" + str(i)] = str(Dato[3][3])  # Más datos
                sheet["W" + str(i)] = str(Dato[3][4])  # Más datos
                sheet["I" + str(i)] = str(Dato[4][1])  # Más datos
                sheet["X" + str(i)] = str(Dato[4][2])  # Más datos
                sheet["Y" + str(i)] = str(Dato[4][3])  # Más datos
                sheet["Z" + str(i)] = str(Dato[4][4])  # Más datos
                sheet["J" + str(i)] = str(Dato[5][1])  # Más datos
                sheet["AA" + str(i)] = str(Dato[5][2])  # Más datos
                sheet["AB" + str(i)] = str(Dato[5][3])  # Más datos
                sheet["AC" + str(i)] = str(Dato[5][4])  # Más datos
                sheet["K" + str(i)] = str(Dato[6][1])  # Más datos
                sheet["AD" + str(i)] = str(Dato[6][2])  # Más datos
                sheet["AE" + str(i)] = str(Dato[6][3])  # Más datos
                sheet["AF" + str(i)] = str(Dato[6][4])  # Más datos
                for j in range(1, 8):  # Ajusta según la cantidad de Dato
                        col_letter = chr(70 + j)  # Calcula la letra de la columna
                        if sheet[col_letter + str(i)].value is None and len(Dato) > j:
                            sheet[col_letter + str(i)] = str(Dato[j])  # Asigna el valor
            else:
                return False
    else:
        for i in range(4, x + 4):
            Dato = verificar_asistencias(code[(i - 4)],periodo)
            if Dato is not None:
                sheet["A" + str(i)] = str(Dato[0][0])  # Código de empleado
                sheet["B" + str(i)] = str(nombres[i - 4])  # Nombre completo del empleado
                sheet["C" + str(i)] = str(ape[i - 4])  # Nombre completo del empleado
                sheet["D" + str(i)] = str(ape2[i - 4])  # Nombre completo del empleado
                # Solo asignar si hay datos en Dato[0] y Dato[0][1]
                sheet["E" + str(i)] = 1 if Dato[0][1] > 0 else 0  # Dias_Asistencia (0 o 1)
                sheet["U" + str(i)] = str(Dato[0][2])  # Horas_Extra
                sheet["V" + str(i)] = str(Dato[0][3])  # Turnos_Extras
                sheet["W" + str(i)] = str(Dato[0][4])  # Descansos_Trabajados
                # Asegúrate de verificar la longitud de Dato[1] y demás
                sheet["F" + str(i)] = str(Dato[1][1])  # Datos de asistencia
                sheet["X" + str(i)] = str(Dato[1][2])  # Más datos
                sheet["Y" + str(i)] = str(Dato[1][3])  # Más datos
                sheet["Z" + str(i)] = str(Dato[1][4])  # Más datos
                sheet["G" + str(i)] = str(Dato[2][1])  # Datos adicionales
                sheet["AA" + str(i)] = str(Dato[2][2])  # Más datos
                sheet["AB" + str(i)] = str(Dato[2][3])  # Más datos
                sheet["AC" + str(i)] = str(Dato[2][4])  # Más datos
                sheet["H" + str(i)] = str(Dato[3][1])  # Más datos
                sheet["AD" + str(i)] = str(Dato[3][2])  # Más datos
                sheet["AE" + str(i)] = str(Dato[3][3])  # Más datos
                sheet["AF" + str(i)] = str(Dato[3][4])  # Más datos
                sheet["I" + str(i)] = str(Dato[4][1])  # Más datos
                sheet["AG" + str(i)] = str(Dato[4][2])  # Más datos
                sheet["AH" + str(i)] = str(Dato[4][3])  # Más datos
                sheet["AI" + str(i)] = str(Dato[4][4])  # Más datos
                sheet["J" + str(i)] = str(Dato[5][1])  # Más datos
                sheet["AJ" + str(i)] = str(Dato[5][2])  # Más datos
                sheet["AK" + str(i)] = str(Dato[5][3])  # Más datos
                sheet["AL" + str(i)] = str(Dato[5][4])  # Más datos
                sheet["K" + str(i)] = str(Dato[6][1])  # Más datos
                sheet["AM" + str(i)] = str(Dato[6][2])  # Más datos
                sheet["AN" + str(i)] = str(Dato[6][3])  # Más datos
                sheet["AO" + str(i)] = str(Dato[6][4])  # Más datos
                sheet["L" + str(i)] = str(Dato[7][1])  # Más datos
                sheet["AP" + str(i)] = str(Dato[7][2])  # Más datos
                sheet["AQ" + str(i)] = str(Dato[7][3])  # Más datos
                sheet["AR" + str(i)] = str(Dato[7][4])  # Más datos
                sheet["M" + str(i)] = str(Dato[8][1])  # Más datos
                sheet["AS" + str(i)] = str(Dato[8][2])  # Más datos
                sheet["AT" + str(i)] = str(Dato[8][3])  # Más datos
                sheet["AU" + str(i)] = str(Dato[8][4])  # Más datos
                sheet["N" + str(i)] = str(Dato[9][1])  # Más datos
                sheet["AV" + str(i)] = str(Dato[9][2])  # Más datos
                sheet["AW" + str(i)] = str(Dato[9][3])  # Más datos
                sheet["AX" + str(i)] = str(Dato[9][4])  # Más datos
                sheet["O" + str(i)] = str(Dato[10][1])  # Más datos
                sheet["AY" + str(i)] = str(Dato[10][2])  # Más datos
                sheet["AZ" + str(i)] = str(Dato[10][3])  # Más datos
                sheet["BA" + str(i)] = str(Dato[10][4])  # Más datos
                sheet["P" + str(i)] = str(Dato[11][1])  # Más datos
                sheet["BB" + str(i)] = str(Dato[11][2])  # Más datos
                sheet["BC" + str(i)] = str(Dato[11][3])  # Más datos
                sheet["BD" + str(i)] = str(Dato[11][4])  # Más datos
                sheet["Q" + str(i)] = str(Dato[12][1])  # Más datos
                sheet["BE" + str(i)] = str(Dato[12][2])  # Más datos
                sheet["BF" + str(i)] = str(Dato[12][3])  # Más datos
                sheet["BG" + str(i)] = str(Dato[12][4])  # Más datos
                sheet["R" + str(i)] = str(Dato[13][1])  # Más datos
                sheet["BH" + str(i)] = str(Dato[13][2])  # Más datos
                sheet["BI" + str(i)] = str(Dato[13][3])  # Más datos
                sheet["BJ" + str(i)] = str(Dato[13][4])  # Más datos
                sheet["S" + str(i)] = str(Dato[14][1])  # Más datos
                sheet["BK" + str(i)] = str(Dato[14][2])  # Más datos
                sheet["BL" + str(i)] = str(Dato[14][3])  # Más datos
                sheet["BM" + str(i)] = str(Dato[14][4])  # Más datos
                if len(Dato)>15:
                    sheet["T" + str(i)] = str(Dato[15][1])  # Más datos
                    sheet["BN" + str(i)] = str(Dato[15][2])  # Más datos
                    sheet["BO" + str(i)] = str(Dato[15][3])  # Más datos
                    sheet["BP" + str(i)] = str(Dato[15][4])  # Más datos
            else:
                return False
    # Ajustar el ancho de las columnas automáticamente
    # Guardar el archivo de Excel actualizado
    workbook.save(ruta+"\\Formatollenado.xlsx")
    workbook.close()
    periodo=periodo.replace(" ", "_")
    periodo=periodo.replace("/", "-")
    if not os.path.exists(os.path.dirname(sys.executable)+f"\\Reportes\\{depar}"):
        os.mkdir(os.path.dirname(sys.executable)+f"\\Reportes\\{depar}")
    if not os.path.exists(os.path.dirname(sys.executable)+f"\\Reportes\\{depar}\\{periodo}"):
        os.mkdir(os.path.dirname(sys.executable)+f"\\Reportes\\{depar}\\{periodo}")
    # Convertir a PDF usando pandas y matplotlib
    ruta_excel = ruta+"\\Formatollenado.xlsx"
    pdf_output = os.path.dirname(sys.executable)+f"\\Reportes\\{depar}\\{periodo}\\Reporte_"+periodo+"_proyecto_"+depar+".pdf"
    excel_to_pdf(ruta_excel, pdf_output,tipo,depar)
    return True   
def excel_to_pdf(excel_file, pdf_file,tipo,depar):
    # Initialize Excel application (headless)
    ruta= os.path.dirname(__file__)
    try:
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = True  # Keep Excel hidden
        # Open the workbook
        workbook = excel.Workbooks.Open(excel_file)
        # Save as PDF
        workbook.Sheets(1).Columns.AutoFit()
        workbook.ExportAsFixedFormat(0, ruta+"\\_"+depar+"pdf_file.pdf")
        # Close the workbook and Excel application
    except Exception as e:
        logging.error(f"Error en excel: {e}")
    finally:
        # Cerrar workbook y Excel
        if workbook:
            workbook.Close(SaveChanges=True)
        if excel:
            del excel
    # Clean up resources
    def encrypt_pdf(input_pdf, output_pdf, user_password, owner_password):
        # Encrypt the PDF using pikepdf
        with pikepdf.open(input_pdf,allow_overwriting_input=True) as pdf:
            pdf.save(
                output_pdf,
                encryption=pikepdf.Encryption(
                    user=user_password,
                    owner=owner_password,
                    allow=pikepdf.Permissions(extract=False, print_lowres=False, modify_annotation=False,modify_assembly=False,modify_form=False,modify_other=False)
                )
            )
        if os.path.exists(input_pdf):
            try:
                os.remove(input_pdf)  # Eliminar archivo existente
            except PermissionError as e:
                raise RuntimeError(f"No se pudo eliminar el archivo {input_pdf}: {e}")
    # Set file paths and passwords
    original_pdf = ruta+"\\_"+depar+"pdf_file.pdf"
    if tipo=="Sindicato":
        encrypted_pdf =pdf_file
    else:
        encrypted_pdf =pdf_file
    user_password = "userpass123"
    owner_password = "ownerpass456"
    if os.path.exists(encrypted_pdf):
        try:
            os.remove(encrypted_pdf)  # Eliminar archivo existente
        except PermissionError as e:
            raise RuntimeError(f"No se pudo eliminar el archivo {encrypted_pdf}: {e}")
    # Create the PDF and then encrypt it
    encrypt_pdf(original_pdf, encrypted_pdf, user_password, owner_password)

def almacenar_fecha(dia_semana, var_checkbox, dia_texto, nomina,fechas_seleccionadas,codigo,index):
    var=Variables()
    year,month,day,___,__,_=var.obtener_fechas()
    dias_festivos=calcular_fechas()
    # Obtener el día de la semana actual
    hoy = datetime.now()
    # Calcular la fecha del día seleccionado
    if nomina == "2" and day <=15 and codigo in str(index):
        fecha = hoy - timedelta(days=day) + timedelta(days=dia_semana+1)
    elif nomina=="2" and day>15 and codigo in str(index):
        fecha=datetime(year,month,16)+timedelta(days=dia_semana)
    elif nomina=="1":
        fecha = hoy - timedelta(days=hoy.weekday()) + timedelta(days=dia_semana+0)
    # Formatear la fecha como DD/MM/AAAA
    fecha_formateada = fecha.strftime('%d/%m/%Y')
    if var_checkbox==1 and codigo in str(index):
        # Si el checkbox está marcado, agregamos la fecha
        return fecha_formateada
    
def agregar_dato(dias, comentario, periodo, aprovacion, codigoEmpleado, HE, DF, TE, DT,descanso,nomina):
    multireg=False
    vari=Variables()
    year,month,day,dias_festivos,fechas_seleccionadas,_=vari.obtener_fechas()
    dias_festivos=calcular_fechas()
    db = ConexionBD(host="148.200.128.13", database="BdTrabajadTemporal",user="andres",password="Andr3s2024")
    global buttonClicked
    if nomina=="2":
        for i, ((dia,index), var) in enumerate(dias.items()):
            if str(codigoEmpleado) in index:
                fechas_seleccionadas[dia,str(index[1]).strip(),codigoEmpleado]=almacenar_fecha(index[1], var, dia, nomina,fechas_seleccionadas,codigoEmpleado,index)  # Guarda las fechas seleccionadas
    else:
        for i, ((dia,index), var) in enumerate(dias.items()):
            if str(codigoEmpleado) in index:
                fechas_seleccionadas[dia,str(index[1]).strip(),codigoEmpleado]=almacenar_fecha(index[1], var, dia, nomina,fechas_seleccionadas,codigoEmpleado,index)  # Guarda las fechas seleccionadas
    try:
        # Conectar a la base de datos
        db.conectar()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        resultado = db.ejecutar_consulta("SELECT COUNT(*) FROM Prenomina WHERE Codigo_Empleado = ? AND periodo= ?", (codigoEmpleado,periodo))
        if "0" in str(resultado):
            # Si no existe, insertar un nuevo registro
            for i, ((dia,index), var) in enumerate(dias.items()):
                    if str(codigoEmpleado) in index:
                        key=(codigoEmpleado,index[1])
                        # Iterar sobre los días festivos
                        for dia_festivo in dias_festivos:
                            if str(fechas_seleccionadas.get((dia,str(index[1]).strip(),codigoEmpleado), 0)) == dia_festivo.strftime('%d/%m/%Y'):
                                DF = "1"  
                        if dia in str(descanso) and var:
                            DT[index[1]]="1"
                        db.ejecutar_consulta(
                            """
                            INSERT INTO Prenomina 
                            (Codigo_Empleado, TipoCobro, Dia_Semana, Dia_Asistencia, Horas_Extra, Dias_Festivos, Turnos_Extras, Descansos_Trabajados, Periodo, Aprobacion) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                codigoEmpleado,
                                nomina,           # Tipo de cobro
                                dia,                   # Día de la semana
                                str(fechas_seleccionadas.get((dia,str(index[1]).strip(),codigoEmpleado), 0)),  # Asistencia del día
                                HE.get((key), "0"),         # Horas extra del día i
                                DF,                    # Días festivos (DF) es fijo
                                TE.get((key), "0"),         # Turnos extras del día i
                                DT.get((key), "0"),         # Descansos trabajados del día i
                                periodo,               # Periodo actual
                                aprovacion             # Aprobación
                            ),commit=True
                        )    
            db.commit()
            if multireg==True and buttonClicked==True:
                buttonClicked=False
        else:
            # Si ya existe, realizar un update
            for i, ((dia,index), var) in enumerate(dias.items()):
                    if str(codigoEmpleado) in index:
                        key=(codigoEmpleado,index[1])
                        # Iterar sobre los días festivos
                        for dia_festivo in dias_festivos:
                            if str(fechas_seleccionadas.get((dia,str(index[1]).strip(),codigoEmpleado), 0)) == dia_festivo.strftime('%d/%m/%Y'):
                                DF = "1"  
                        if dia in str(descanso) and var:
                            DT[index[1]]="1"
                        db.ejecutar_consulta(
                            """
                            UPDATE Prenomina SET 
                            Dia_Asistencia = ?,
                            Horas_Extra = ?, Dias_Festivos = ?, Turnos_Extras = ?, Descansos_Trabajados = ?, Aprobacion = ?
                            WHERE Codigo_Empleado = ? AND Dia_Semana = ? AND Periodo = ?
                            """,
                            (
                                str(fechas_seleccionadas.get((dia,str(index[1]).strip(),codigoEmpleado), 0)),  # Asistencia del día
                                HE.get((key), "0"),         # Horas extra del día i
                                DF,                    # Días festivos (DF) es fijo
                                TE.get((key), "0"),         # Turnos extras del día i
                                DT.get((key), "0"),         # Descansos trabajados del día i
                                aprovacion,            # Aprobación
                                codigoEmpleado,        # Código del empleado
                                dia,                   # Día de la semana
                                periodo                # Periodo actual
                            ),commit=True
                        )    
            db.commit()
            if multireg==True and buttonClicked==True:
                buttonClicked=False
        db.cerrar()
    except Exception as e:
        logging.error("Error", f"No se pudo añadir o actualizar el dato: {e}")
# Specify the Excel and PDF file paths
class ConexionBD:
    def __init__(self, host, database, user, password, driver="ODBC Driver 18 for SQL Server"):
        """
        Inicializa una conexión con la base de datos.

        :host: Nombre del host o dirección del servidor.
        :database: Nombre de la base de datos.
        :user: Usuario de la base de datos.
        :password: Contraseña del usuario.
        :driver: Controlador ODBC (por defecto SQL Server).
        """
        self.host = host
        self.database = database
        self.user = user
        self.password = password
        self.driver = driver
        self.conexion = None
        self.cursor = None

    def conectar(self):
        """
        Establece la conexión a la base de datos.
        """
        
        try:
            self.conexion = pyodbc.connect(
                f"DRIVER={{{self.driver}}};"
                f"SERVER={self.host};"
                f"DATABASE={self.database};"
                f"UID={self.user};"  # Usuario
                f"PWD={self.password};"  # Contraseña
                f"TrustServerCertificate=YES;"
            )
            self.cursor = self.conexion.cursor()
        except Exception as e:
            logging.error((f"Error al conectar con la base de datos {self.database}: {e}"))
            raise ConnectionError(f"Error al conectar con la base de datos {self.database}: {e}")

    def ejecutar_consulta(self, query, parametros=None,commit=False):
        """
        Ejecuta una consulta en la base de datos.

        :param query: Consulta SQL a ejecutar.
        :param parametros: Parámetros para la consulta (opcional).
        :return: Lista con los resultados de la consulta.
        """
        try:
            if not self.cursor:
                logging.error("No hay una conexión activa. Llama a 'conectar' primero.")
                raise ConnectionError("No hay una conexión activa. Llama a 'conectar' primero.")
            
            self.cursor.execute(query, parametros or ())
            
            # Si es una consulta de lectura (SELECT), devuelve los resultados
            if query.strip().upper().startswith("SELECT"):
                return self.cursor.fetchall()
            
            # Si es una consulta de escritura, confirma los cambios si se solicita
            if commit:
                self.conexion.commit()
                return None
        except pyodbc.Error as e:
            logging.error(f"Error de base de datos: {e}")
            raise RuntimeError(f"Error al ejecutar la consulta: {e}")

    def commit(self):
        """
        Aplica los cambios a la base de datos.
        """
        if self.conexion:
            self.conexion.commit()

    def cerrar(self):
        """
        Cierra la conexión y libera los recursos.
        """
        if self.cursor:
            self.cursor.close()
        if self.conexion:
            self.conexion.close()
class AnimatedApp(ft.UserControl):
    def __init__(self):
        super().__init__()
         # Crear el Dropdown de departamentos vacío inicialmente
        var=Variables()
        periodos,dias,asis,multireg,todos,excel_file,pdf_file,HE_entries,DT_entries,TE_entries=var.obtener_todos()
        self.HE_entries=HE_entries
        self.TE_entries=TE_entries
        self.DT_entries=DT_entries
        self.periodos=periodos
        self.is_syncing_scroll = False
        self.scroll_timer = None    
        self.asis={}
        self.valid={}
        self.dropdown_departamentos = ft.Dropdown(
            width=150,
            height=40,
            bgcolor=ft.colors.GREY_300,
            color=ft.colors.BLACK,
            hint_text="Seleccione...",
            padding=ft.padding.only(left=10, right=10),
            disabled=True,
            on_change=lambda e:self.cambio_departamentos_search(self.periodos,dias,asis,multireg,todos,excel_file,pdf_file,HE_entries,DT_entries,TE_entries,e)
        )
        self.dropdown_departamentos.options=""
        self.add_project_button = ft.ElevatedButton(
            text="Añadir Proyectos",
            disabled=False,
            bgcolor=ft.colors.BLUE_800,
            color=ft.colors.WHITE,
            on_click=self.abrir_ventana_departamentos  # Cambiamos el evento a una nueva función
        )
        self.sub_project_button = ft.ElevatedButton(
            text="Eliminar Proyectos",
            disabled=False,
            bgcolor=ft.colors.BLUE_800,
            color=ft.colors.WHITE,
            on_click=self.abrir_ventana_departamentos2  # Cambiamos el evento a una nueva función
        )
        self.confirmar = False
        self.add_enviar_todos = ft.ElevatedButton(
            text="Añadir Todos",
            disabled=True,visible=False,
            bgcolor=ft.colors.BLUE_600,
            color=ft.colors.WHITE,
            on_click=lambda e: self.abrirconf(e)  # Cambiamos el evento a una nueva función
        )
        self.generar_repote_button=ft.ElevatedButton(
                                        text="Generar Reporte",
                                        bgcolor=ft.colors.BLUE_600,
                                        color=ft.colors.WHITE,
                                        disabled=True,visible=False,
                                        on_click=lambda e:self.send_data(tipo_d,tipo_e,"")  # Función que manejará el evento del botón
                                    )
        
        # Crear el diálogo modal para mostrar los departamentos
        self.dialog_departamentos = ft.AlertDialog(
            modal=True,
            title=ft.Text("Seleccione Departamentos"),
            content=self.crear_contenido_dialogo(),
            actions=[
                ft.ElevatedButton(
                    text="Cerrar",
                    on_click=lambda e: self.cerrar_ventana_departamentos()  # Llama a la función de cerrar
                )
            ]
        )
        self.dialog_departamentos2 = ft.AlertDialog(
            modal=True,
            title=ft.Text("Seleccione Departamentos"),
            content=self.crear_contenido_dialogo2(),
            actions=[
                ft.ElevatedButton(
                    text="Cerrar",
                    on_click=lambda e: self.cerrar_ventana_departamentos2()  # Llama a la función de cerrar
                )
            ]
        )
        self.dropdown_tipo_empleado = ft.Dropdown(
            width=150,
            height=40,
            bgcolor=ft.colors.GREY_300,
            color=ft.colors.BLACK,
            hint_text="Seleccione...",
            options=[
                ft.dropdown.Option("Confianza"),
                ft.dropdown.Option("Sindicato")
            ],
            padding=ft.padding.only(left=10, right=10),
            on_change=self.tipo_empleado_cambiado  # Llama a la función cuando cambia el valor
        )
        self.add_project_button.visible= verificar_Permisos2(nombre)
        self.sub_project_button.visible= verificar_Permisos2(nombre)
        self.generar_repote_button.visible= verificar_Permisos2(nombre)
        self.generar_repote_button.disabled=not verificar_Permisos2(nombre)
        # Llenar el Dropdown de departamentos
        self.llenar_departamentos()
        empleados=obtener_empleados(2301,"Confianza")
        # Variables de color de ejemplo; reemplázalas según sea necesario
        self.color_title = ft.colors.BLUE_ACCENT_700
        self.color_container = ft.colors.LIGHT_BLUE_100
        self.title_color = ft.colors.BLACK  # Color del título

        # Crear el contenedor de la imagen con un pequeño padding a la izquierda
        self.image = ft.Container(
            content=ft.Image(
                src="https://jumapam.gob.mx/images/JPG/jumapam.jpg",
                width=40,
                height=40,
                fit=ft.ImageFit.CONTAIN
            ),
            padding=ft.padding.only(left=10)  # Espacio a la izquierda
        )

        # Configuración del título
        self.title_text = ft.Text("Jumapam", size=30, color=self.title_color, weight=ft.FontWeight.BOLD)

        # Contenedor vacío para crear un espacio adicional
        self.spacing_container = ft.Container(
            width=200  # Ancho del espacio a la derecha del título
        )
        # Contenedor donde va el periodo, con padding a la izquierda
        self.white_container = ft.Container(
            bgcolor=ft.colors.WHITE,  # Color de fondo blanco
            width=400,  # Ancho del contenedor
            height=30,  # Altura del contenedor
            alignment=ft.alignment.center,  # Alinear contenido al centro
            border_radius=5,
        )
        # Contenedor con imagen, título y botón en una fila
        self.frame_title = ft.Container(
            expand=False,
            bgcolor=self.color_title,
            border_radius=10,
            alignment=ft.alignment.center_left,
            content=ft.Column(controls=[ft.Row(
                controls=[
                    self.image,
                    self.title_text,
                    self.spacing_container,  # Espacio antes del contenedor blanco
                    self.white_container,
                ],
                vertical_alignment=ft.CrossAxisAlignment.CENTER
            ),ft.Row(  # Usar Row para alinear horizontalmente
                                        controls=[
                                            self.dropdown_tipo_empleado,
                                            ft.Container(
                                                width=20  # Espacio entre los Dropdowns
                                            ),
                                            ft.Text("Proyecto:", size=16, color="BLACK"),  # Texto descriptivo
                                            self.dropdown_departamentos,
                                            ft.TextField(hint_text="Buscar Empleado por Proyecto...",value="",bgcolor="white",on_submit=lambda e:self.cambio_departamentos_search(self.periodos,dias,asis,multireg,todos,excel_file,pdf_file,HE_entries,DT_entries,TE_entries,e)),
                                            self.generar_repote_button,
                                            self.add_project_button,
                                            self.sub_project_button,
                                            self.add_enviar_todos
                                        ],alignment=ft.MainAxisAlignment.SPACE_EVENLY,
                                    )])
        )
        self.contenedor_empleados=ft.Container( # Fondo azul claro para la fila
        padding=ft.padding.only(top=20),expand=True,height=560,
                        bgcolor=self.color_container,  # Altura del contenedor
            alignment=ft.alignment.top_left,  # Alinear contenido al centro
            border_radius=5,content=ft.Column(scroll=ft.ScrollMode.ALWAYS))
        self.contenedor_empleados2=ft.Container(padding=ft.padding.only(top=20),expand=False,height=560,
                        bgcolor=self.color_container,
            alignment=ft.alignment.top_center,  # Alinear contenido al centro
            border_radius=5,content=ft.Column(scroll=ft.ScrollMode.ALWAYS))
        tipo_d = self.dropdown_departamentos.value
        tipo_e = self.dropdown_tipo_empleado.value
        diasq=obtener_quincenas()
        dias=sortdias(diasq)
        # Asignar las opciones generadas al Dropdown de departamentos
        emple=obtener_empleados(tipo_d,tipo_e)
        # Contenedor que contiene los Dropdowns, la tabla y el nuevo botón
        # Agregar los contenedores a la página
        self.controls = [
            ft.Column(
                expand=True,
                controls=[
                    self.frame_title,
                    ft.Row(controls=[self.contenedor_empleados2,self.contenedor_empleados]),
                ],scroll=ft.ScrollMode.ALWAYS
            )
        ]
        
    def depar(self, e):
        # Obtener el usuario seleccionado
        usuario_seleccionado = self.dropdown_usuarios2.value
        print(usuario_seleccionado)
        # Obtener departamentos asignados
        id=self.id[usuario_seleccionado]
        depars = obtener_ASIGNED_departamentos(id)
        # Actualizar visibilidad de los checkboxes
        for i, checkbox in enumerate(self.checkboxes_departamentos2):
            if checkbox.label in depars:
                checkbox.visible = True  # Mostrar si pertenece a los asignados
            else:
                checkbox.visible = False  # Ocultar si no pertenece
        self.page.update()  # Actualizar la página para reflejar cambios
    def crear_contenido_dialogo2(self):
        # Lista de departamentos (puedes personalizar esta lista)
        depars=[]
        depars=obtener_ALL_departamentos()
        # Crear un campo de búsqueda para filtrar departamentos
        self.dropdown_usuarios2=ft.Dropdown(
            width=150,
            height=40,
            bgcolor=ft.colors.GREY_300,
            color=ft.colors.BLACK,
            hint_text="Seleccione...",
            padding=ft.padding.only(left=10, right=10),
            on_change=self.depar
        )
        id,usuarios=GET_USER()
        self.id=id
        opciones_usuarios = [ft.dropdown.Option(user) for user in usuarios]
        self.dropdown_usuarios2.options=opciones_usuarios
        # Crear checkboxes para cada departamento
        self.checkboxes_departamentos2 = [
            ft.Checkbox(label=departamento,visible=False) for departamento in depars
        ]
        checkbox_container = ft.Container(
            content=ft.Column(controls=self.checkboxes_departamentos2, spacing=5,height=300,scroll="auto")
        )
        # Botón para mostrar los departamentos seleccionados
        self.show_selected_button2 = ft.ElevatedButton(
            text="Eliminar Proyectos",
            on_click=self.mostrar_departamentos_seleccionados2
        )

        # Retornar el contenido del diálogo en un contenedor Column
        return ft.Column(
            controls=[
                self.dropdown_usuarios2,
                checkbox_container,
                self.show_selected_button2
            ]
        )
    def activate(self,e):
        usuario_seleccionado = self.dropdown_usuarios.value
        print(usuario_seleccionado)
        # Obtener departamentos asignados
        id=self.id[usuario_seleccionado]
        depars = obtener_ASIGNED_departamentos(id)
        print(depars)
        for i, checkbox in enumerate(self.checkboxes_departamentos):
            if checkbox.label in depars:
                checkbox.visible = False  # Mostrar si pertenece a los asignados
            else:
                checkbox.visible = True  # Ocultar si no pertenece
        self.page.update()
    def crear_contenido_dialogo(self):
        # Lista de departamentos (puedes personalizar esta lista)
        depars=[]
        depars=obtener_ALL_departamentos()
        # Crear un campo de búsqueda para filtrar departamentos
        self.search_field = ft.TextField(
            hint_text="Buscar Proyecto...",
            on_change=self.filtrar_departamentos
        )
        self.dropdown_usuarios=ft.Dropdown(
            width=150,
            height=40,
            bgcolor=ft.colors.GREY_300,
            color=ft.colors.BLACK,
            hint_text="Seleccione...",
            padding=ft.padding.only(left=10, right=10),
            on_change=self.activate
        )
        self.dropdown_usuarios.options=""
        id,usuarios=GET_USER()
        self.id=id
        opciones_usuarios = [ft.dropdown.Option(user) for user in usuarios]
        self.dropdown_usuarios.options=opciones_usuarios
        # Crear checkboxes para cada departamento
        self.checkboxes_departamentos = [
            ft.Checkbox(label=departamento,visible=False) for departamento in depars
        ]
        checkbox_container = ft.Container(
            content=ft.Column(controls=self.checkboxes_departamentos, spacing=5,height=300,scroll="auto")
        )
        # Botón para mostrar los departamentos seleccionados
        self.show_selected_button = ft.ElevatedButton(
            text="Añadir Proyectos",
            on_click=self.mostrar_departamentos_seleccionados
        )
        # Retornar el contenido del diálogo en un contenedor Column
        return ft.Column(
            controls=[
                self.dropdown_usuarios,
                self.search_field,
                checkbox_container,
                self.show_selected_button
            ]
        )
    def confirmacion(self, e, respuesta):
        self.confirmar = respuesta
        self.ventanaconf.open = False
        self.page.overlay.remove(self.ventanaconf)  
        self.page.update()
        self.añadir_todos(self.asis,self.HE_entries,self.DT_entries,self.TE_entries)

    def añadir_todos(self, asis, HE_entries, DT_entries, TE_entries):
        try:
            todos={}
            periodos=""
            global multireg
            multireg=False
            tipo_dep = self.dropdown_departamentos.value
            tipo_empleado = self.dropdown_tipo_empleado.value
            empleados=obtener_empleados(tipo_dep,tipo_empleado)
            periodos=self.tipo_empleado_cambiado()
            print(self.confirmar)
            if not periodos:
                logging.error("El valor de periodos no es válido.")
                return
            if not empleados:
                logging.error("No se encontraron empleados para procesar.")
                return
            if self.confirmar:
                [agregar_dato(
                                                                    {dia: var for dia, var in asis.items()},
                                                                    "",
                                                                    periodos,
                                                                    self.valid[empleado[1][1],empleado[1][0]],
                                                                    empleado[1][0],
                                                                    HE_entries,
                                                                    "",
                                                                    DT_entries,
                                                                    TE_entries,
                                                                    "Domingo",
                                                                    empleado[1][1]
                                                                )for empleado in enumerate(empleados)]
                alert=ft.AlertDialog(
                title=ft.Text("Todos los datos Datos añadidos correctamente."),
                )
                self.page.open(alert)
            else:
                print("ALERTA")
                alert=ft.AlertDialog(
                title=ft.Text("Operacion Cancelada."),
                )
                self.page.open(alert)
        except Exception as e:
            logging.error(f"Error en mostrar_departamentos_seleccionados: {e}")
    # Función para abrir el diálogo de departamentos
    def abrirconf(self,e):
        self.ventanaconf = ft.AlertDialog(
                modal=True,
                title=ft.Text("Desea Añadir Multiples Entradas"),
                content=ft.Text("Confirme la informacion antes de proceder"),
                actions=[
                    ft.ElevatedButton(
                        text="Confirmar",
                        on_click=lambda e: self.confirmacion(e,True)
                    ),
                    ft.ElevatedButton(
                        text="Cerrar",
                        on_click=lambda e: self.confirmacion(e,False)
                    )
                ]
            )
        self.page.overlay.append(self.ventanaconf)
        self.ventanaconf.open = True
        self.page.update()
    def abrir_ventana_departamentos2(self, e):
        # Agregar el diálogo a la lista de overlays y abrirlo
        self.page.overlay.append(self.dialog_departamentos2)
        self.dialog_departamentos2.open = True
        self.page.update()
    def abrir_ventana_departamentos(self, e):
        # Agregar el diálogo a la lista de overlays y abrirlo
        self.page.overlay.append(self.dialog_departamentos)
        self.dialog_departamentos.open = True
        self.page.update()
    def cerrar_ventana_departamentos2(self):
        self.dialog_departamentos2.open = False
        self.page.update()
        self.page.overlay.remove(self.dialog_departamentos2)
    def cerrar_ventana_departamentos(self):
        self.dialog_departamentos.open = False
        self.page.update()
        self.page.overlay.remove(self.dialog_departamentos)

    # Función para filtrar departamentos en base a la búsqueda
    def filtrar_departamentos(self, e):
        texto_busqueda = self.search_field.value.lower()
        for checkbox in self.checkboxes_departamentos:
            checkbox.visible = texto_busqueda in checkbox.label.lower()
        self.page.update()

    # Función para mostrar los departamentos seleccionados
    def mostrar_departamentos_seleccionados2(self, e):
        try:
            if self.dropdown_usuarios2.value != "":
                seleccionados = [checkbox.label for checkbox in self.checkboxes_departamentos2 if checkbox.value]
                x=self.id[self.dropdown_usuarios2.value]
                usuario=DEL_USER(x,seleccionados)
                departamentos.clear()
                obtener_departamentos(usuario[0])
                opciones_departamentos = [ft.dropdown.Option(depto) for depto in departamentos]
                # Asignar las opciones generadas al Dropdown de departamentos
                self.dropdown_departamentos.options = opciones_departamentos
                self.cerrar_ventana_departamentos2()
                for checkbox in self.checkboxes_departamentos2:
                    if checkbox.value:
                        checkbox.value=False
                        checkbox.visible=False
                self.update()
                alert=ft.AlertDialog(
                title=ft.Text("Eliminado con exito."),
                )
                self.page.open(alert)
            else:
                alert=ft.AlertDialog(
                title=ft.Text("Usuario no Seleccionado."),
                )
                self.page.open(alert)
        except Exception as e:
            logging.error(f"Error en mostrar_departamentos_seleccionados: {e}")
    def mostrar_departamentos_seleccionados(self, e):
        try:
            if self.dropdown_usuarios.value != "":
                seleccionados = [checkbox.label for checkbox in self.checkboxes_departamentos if checkbox.value]
                x=self.id[self.dropdown_usuarios.value]
                usuario=UPDATE_USER(x,seleccionados)
                departamentos.clear()
                obtener_departamentos(usuario[0])
                opciones_departamentos = [ft.dropdown.Option(depto) for depto in departamentos]
                # Asignar las opciones generadas al Dropdown de departamentos
                self.dropdown_departamentos.options = opciones_departamentos
                self.cerrar_ventana_departamentos()
                for checkbox in self.checkboxes_departamentos:
                    if checkbox.value:
                        checkbox.visible=False
                        checkbox.value=False
                self.update()
                alert=ft.AlertDialog(
                title=ft.Text("Añadido con exito."),
                )
                self.page.open(alert)
            else:
                alert=ft.AlertDialog(
                title=ft.Text("Usuario no Seleccionado."),
                )
                self.page.open(alert)
        except Exception as e:
            logging.error(f"Error en mostrar_departamentos_seleccionados: {e}")
    def llenar_departamentos(self):
        # Crear las opciones para el Dropdown de departamentos
        opciones_departamentos = [ft.dropdown.Option(depto) for depto in departamentos]
        # Asignar las opciones generadas al Dropdown de departamentos
        self.dropdown_departamentos.options = opciones_departamentos
    def datos(self,e,asis,HE_entries,DT_entries,TE_entries):
        tipo_dep = self.dropdown_departamentos.value
        tipo_empleado = self.dropdown_tipo_empleado.value
        empleados=obtener_empleados(tipo_dep,tipo_empleado)
        periodos=self.tipo_empleado_cambiado()
        [agregar_dato(
                                                            {dia: var for dia, var in self.asis.items()},
                                                            "",
                                                            periodos,
                                                            self.valid[e.control.key[1],e.control.key[0]],
                                                            e.control.key[0],
                                                            HE_entries,
                                                            "",
                                                            DT_entries,
                                                            TE_entries,
                                                            "Domingo",
                                                            e.control.key[1]
                                                        )]
        alert=ft.AlertDialog(
        title=ft.Text("Datos añadidos correctamente."),
        )
        self.page.open(alert)
    def debounce_scroll(self, scroll_handler, delay=0.1):
        """
        Debounce para los eventos de desplazamiento.
        """
        def wrapper(e):
            if self.scroll_timer:
                self.scroll_timer.cancel()  # Cancelar el temporizador anterior

            # Crear un nuevo temporizador
            self.scroll_timer = threading.Timer(delay, lambda: scroll_handler(e))
            self.scroll_timer.start()

        return wrapper

    def sync_scroll(self,e: ft.ScrollEvent):
        print("scrollCAJAS")
        # Sincronizar scroll del segundo contenedor con el primero
        if not self.is_syncing_scroll:  # Solo sincronizar si no se está ya sincronizando
            self.is_syncing_scroll = True
            self.contenedor_empleados2.content.scroll_to(offset=e.pixels)
            self.is_syncing_scroll = False
    def sync_scroll2(self,e: ft.ScrollEvent):
        print("scrollNOMBRES")
        # Sincronizar scroll del segundo contenedor con el primero
        if not self.is_syncing_scroll:
            self.is_syncing_scroll=True
            self.contenedor_empleados.content.controls[0].scroll_to(offset=e.pixels)
            self.is_syncing_scroll=False
    # Asignar el evento de scroll al primer contenedor
    def move(self,e):
            if e=="Retroceder":
                self.contenedor_empleados.content.scroll_to(offset=0)
            else:
                self.contenedor_empleados.content.scroll_to(offset=-1)
            self.page.update()
    def cambio_departamentos_search(self,periodos,dias,asis,multireg,todos,excel_file,pdf_file,HE_entries,DT_entries,TE_entries, e):
        self.add_enviar_todos.disabled=False
        self.add_enviar_todos.visible=True
        var=Variables()
        usuario=var.obtener_usuario()
        dias_semana,semana=var.obtener_semana()
        def updateHE(self):
            # Agregar el diálogo a la lista de overlays y abrirlo
            HE_entries[self.control.key]=self.control.value
        def updateDT(self):
            # Agregar el diálogo a la lista de overlays y abrirlo
            DT_entries[self.control.key]=self.control.value
        def updateTE(self):
            # Agregar el diálogo a la lista de overlays y abrirlo
            TE_entries[self.control.key]=self.control.value
        # Crear las opciones para el Dropdown de departamentos
        tipo_dep = self.dropdown_departamentos.value
        tipo_empleado = self.dropdown_tipo_empleado.value
        # Asignar las opciones generadas al Dropdown de departamentos
        if e.control.value=="":
            empleados=obtener_empleados_search(tipo_dep,tipo_empleado,"")
        else:
            empleados=obtener_empleados_search(tipo_dep,tipo_empleado,e.control.value)
        if tipo_empleado=="Sindicato":
            self.contenedor_empleados2.content=ft.Column(expand=False)
            filas_empleados=[ft.Container(
                                    bgcolor=ft.colors.BLUE_600,  # Fondo azul claro para la fila
                                    content=ft.Column(
                                        controls=[
                                                    ft.Row(
                                                spacing=10,
                                                controls=[
                                                    ft.Text("ID", size=16, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=70),
                                                    ft.Text("Nombre", size=16, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=200),
                                                    ft.Text("Proyecto", size=16, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=70),
                                                    *[ft.Text(value=semana[i], size=16, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=90) for i in range(7) if tipo_empleado == "Sindicato"],
                                                    ft.Text("Aprobación", size=12, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center",width=70),
                                                ],
                                                alignment=ft.MainAxisAlignment.START
                                            ),
                                            *[ft.Row(
                                                        controls=[
                                                            ft.TextField(value=f"{_[0]}",  height=60, width=70, color="black", bgcolor=ft.colors.BLUE_300, disabled=True, multiline=True, min_lines=1, max_lines=2, text_size=13, border="none", text_align="center"),
                                                            ft.TextField(value=f"{_[2]}  {_[3]}", height=60, width=200, bgcolor=ft.colors.BLUE_300, disabled=True, multiline=True, min_lines=1, max_lines=2, text_size=13, border="none", color="black", text_align="center"),
                                                            ft.TextField(value=f"{_[7]}", height=60, width=70, color="black", bgcolor=ft.colors.BLUE_300, disabled=True, multiline=True, min_lines=1, max_lines=2, text_size=13, border="none", text_align="center"),
                                                            *[ft.Column(
                                                            controls=[
                                                                ft.Row(
                                                                    controls=[
                                                                        ft.TextField(key=(str(_[0]),__),height=30, width=40, color="black", hint_text="HE", bgcolor="white", text_align="center", text_size=10,on_change=updateHE),
                                                                        ft.Checkbox(key=(str(_[0]),__),value=check_dias(semana[__],_[0],self.periodos),height=30, width=40, check_color="black", fill_color="white",on_change=lambda e:self.checkbox_changed(e,asis,dias_semana,_[1])),
                                                                    ],
                                                                ),
                                                                ft.Row(
                                                                    controls=[
                                                                        ft.TextField(key=(str(_[0]),__),height=30, width=40, color="black", hint_text="DT", bgcolor="white", text_align="center", text_size=10,on_change=updateDT),
                                                                        ft.TextField(key=(str(_[0]),__),height=30, width=40, color="black", hint_text="TE", bgcolor="white", text_align="center", text_size=10,on_change=updateTE),
                                                                    ],
                                                                ),
                                                            ]
                                                        ) for __ in range(7)],  # Checkbox para cada día
                                                            ft.Checkbox(key=((str(_[0]),str(_[1]))),value=checar_aprovacion(_[0],self.periodos),disabled=verificar_Permisos(nombre),width=70,on_change=self.Aprovecheck),
                                                            ft.ElevatedButton(key=(str(_[0]),_[1]),text="Añadir",icon=ft.icons.ADD, width=60, height=50, bgcolor=ft.colors.BLUE_900, color=ft.colors.WHITE,on_click=lambda e: self.datos(e,asis,HE_entries,DT_entries,TE_entries))
                                                        ],
                                                    )for index,_ in enumerate(empleados)],
                                        ],
                                    )
                                                )]
            for index,_ in enumerate(empleados):
                for __ in range(7):
                    x=check_dias(semana[__],_[0],self.periodos)
                    self.asis_changed(((str(_[0]),__),semana[__]),x,dias_semana,_[1])
                    self.Aprovecheckbegin((_[0],_[1],False))
            self.contenedor_empleados.content = ft.Column(controls=filas_empleados,scroll=ft.ScrollMode.ALWAYS)
            self.contenedor_empleados.padding=0
            self.white_container.content=ft.Text(self.periodos, color=ft.colors.BLACK)    
        else:
            diasq=obtener_quincenas()
            dias=sortdias(diasq)
            self.contenedor_empleados2.content=ft.Column(  # Fondo azul claro para la fila
                            controls=[
                                ft.Row(controls=[
                                    ft.Text("ID", size=16, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=70),
                                    ft.Text("Nombre", size=16, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=200),
                                    ft.Text("Proyecto", size=16, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=70),
                                ],alignment=ft.MainAxisAlignment.SPACE_EVENLY),
                                *[ft.Row(controls=[
                                    ft.TextField(value=f"{_[0]}", height=70, width=70, color="black", bgcolor=ft.colors.BLUE_300, disabled=True, multiline=True, min_lines=1, max_lines=2, text_size=13, border="none", text_align="center"),
                                    ft.TextField(value=f"{_[2]}  {_[3]}", height=70, width=200, bgcolor=ft.colors.BLUE_300, disabled=True, multiline=True, min_lines=1, max_lines=2, text_size=13, border="none", color="black", text_align="center"),
                                    ft.TextField(value=f"{_[7]}", height=70, width=70, color="black", bgcolor=ft.colors.BLUE_300, disabled=True, multiline=True, min_lines=1, max_lines=2, text_size=13, border="none", text_align="center"),
                                ],alignment=ft.MainAxisAlignment.SPACE_EVENLY) for index, _ in enumerate(empleados)]
                            ],scroll=ft.ScrollMode.ALWAYS,on_scroll=self.debounce_scroll(self.sync_scroll2)
                        )
            filas_empleados = ft.Column(
            controls=[ft.Row(
                    scroll=ft.ScrollMode.ALWAYS,
                    controls=[
                        # Segunda columna (scrollable horizontalmente)
                        ft.Container(content=ft.Row(
                            controls=[
                                ft.Column(
                                    controls=[
                                        ft.Row(controls=[
                                            *[ft.Text(value=dia[0], size=12, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, width=80) for i, (dia, var) in enumerate(dias.items()) if tipo_empleado == "Confianza"],
                                            ft.Text("Aprobación", size=12, weight=ft.FontWeight.BOLD, color=ft.colors.BLACK, text_align="center", width=70),
                                        ]),
                                        *[ft.Row(
                                            controls=[
                                                ft.Row(
                                                    controls=[
                                                        *[ft.Column(
                                                            controls=[
                                                                ft.Row(controls=[
                                                                    ft.TextField(key=(str(_[0]), i), height=30, width=40, color="black", hint_text="HE", bgcolor="white", text_align="center", text_size=10, on_change=updateHE),
                                                                    ft.Checkbox(key=((str(_[0]), i), dias_semana[diasq[i]]), value=check_dias(dias_semana[diasq[i]], _[0], self.periodos), height=30, width=40, check_color="black", fill_color="white", on_change=lambda e: self.checkbox_changed(e, asis, dias_semana, _[1])),
                                                                ], spacing=0),
                                                                ft.Row(controls=[
                                                                    ft.TextField(key=(str(_[0]), i), height=30, width=40, color="black", hint_text="DT", bgcolor="white", text_align="center", text_size=10, on_change=updateDT),
                                                                    ft.TextField(key=(str(_[0]), i), height=30, width=40, color="black", hint_text="TE", bgcolor="white", text_align="center", text_size=10, on_change=updateTE),
                                                                ], spacing=0),
                                                            ]
                                                        ) for i, (dia, var) in enumerate(dias.items())]
                                                    ]
                                                ),
                                                ft.Checkbox(key=((str(_[0]), str(_[1]))), value=checar_aprovacion(_[0],self.periodos),disabled=verificar_Permisos(nombre), width=70, on_change=self.Aprovecheck),
                                                ft.ElevatedButton(key=(str(_[0]), _[1]), text="Añadir", icon=ft.icons.ADD, width=60, height=50, bgcolor=ft.colors.BLUE_900, color=ft.colors.WHITE, on_click=lambda e: self.datos(e, asis, HE_entries, DT_entries, TE_entries))
                                            ]
                                        ) for index, _ in enumerate(empleados)]
                                    ]
                                )
                            ],  # Habilitar desplazamiento horizontal
                        ))
                    ],
                    alignment=ft.MainAxisAlignment.START
                ),

            ],scroll=ft.ScrollMode.ALWAYS,on_scroll=self.debounce_scroll(self.sync_scroll)
        )
            for index,_ in enumerate(empleados):
                for i,(dia,var) in enumerate(dias.items()):
                    x=check_dias(dias_semana[diasq[i]],_[0],self.periodos)
                    self.asis_changed(((str(_[0]),i),dias_semana[diasq[i]]),x,dias_semana,_[1])
                    self.Aprovecheckbegin((_[0],_[1],False))    
            self.contenedor_empleados.content = ft.Row(controls=[filas_empleados],scroll=ft.ScrollMode.ALWAYS)
            self.contenedor_empleados.bgcolor=ft.colors.BLUE_600
            self.contenedor_empleados.padding=ft.padding.only(top=20)
            self.frame_title.content=ft.Container(
            expand=False,
            bgcolor=self.color_title,
            border_radius=10,
            alignment=ft.alignment.center_left,
            content=ft.Column(controls=[ft.Row(
                controls=[
                    self.image,
                    self.title_text,
                    self.spacing_container,  # Espacio antes del contenedor blanco
                    self.white_container,
                ],
                vertical_alignment=ft.CrossAxisAlignment.CENTER
            ),ft.Row(  # Usar Row para alinear horizontalmente
                                        controls=[
                                            self.dropdown_tipo_empleado,
                                            ft.Container(
                                                width=20  # Espacio entre los Dropdowns
                                            ),
                                            ft.Text("Proyecto:", size=20, color="BLACK"),  # Texto descriptivo
                                            self.dropdown_departamentos,
                                            ft.TextField(hint_text="Buscar Empleado por ID...",bgcolor="white",value="",on_submit=lambda e:self.cambio_departamentos_search(self.periodos,dias,asis,multireg,todos,excel_file,pdf_file,HE_entries,DT_entries,TE_entries,e)),
                                            self.generar_repote_button,
                                            self.add_project_button,
                                            self.sub_project_button,
                                            self.add_enviar_todos,
                                            ft.Row(controls=[ft.ElevatedButton(text="Retroceder", icon=ft.icons.ARROW_LEFT, width=60, height=50, bgcolor=ft.colors.BLUE_900, color=ft.colors.WHITE, on_click=lambda e:self.move("Retroceder")),ft.ElevatedButton(text="Avanzar", icon=ft.icons.ARROW_RIGHT, width=60, height=50, bgcolor=ft.colors.BLUE_900, color=ft.colors.WHITE, on_click=lambda e:self.move("Avanzar")),
                                   ],alignment=ft.MainAxisAlignment.SPACE_EVENLY)
                                        ],alignment=ft.MainAxisAlignment.SPACE_EVENLY,
                                    )])
        )
            self.white_container.content=ft.Text(self.periodos, color=ft.colors.BLACK)
        self.update() 
    def Aprovecheck(self,e):
        try:
            self.valid[e.control.key[1],e.control.key[0]]=e.control.value
        except Exception as e:
            logging.error("Error", f"Error al ejecutar el Proceso Almacenado: {e}")
    def Aprovecheckbegin(self,e):
        try:
            self.valid[e[1],e[0]]=e[2]
        except Exception as e:
            logging.error("Error", f"Error al ejecutar el Proceso Almacenado: {e}")
    def checkbox_changed(e,x,asist,dias_semana,nomina):
        if nomina=="1":
            e.asis[dias_semana[x.control.key[1]],(x.control.key[0],x.control.key[1])]=int(x.control.value)
        else:
            e.asis[x.control.key[1],x.control.key[0]]=int(x.control.value)
    def asis_changed(e,x,asist,dias_semana,nomina):
        if nomina=="1":
            if asist==False:
                 e.asis[x[1],x[0]]=0
            else:
                e.asis[x[1],x[0]]=1
        else:
            if asist==False:
                 e.asis[x[1],x[0]]=0
            else:   
                e.asis[x[1],x[0]]=1
    def tipo_empleado_cambiado(self, e=None):
        # Obtiene el valor seleccionado en el primer Dropdown
        self.dropdown_departamentos.disabled=False
        tipo_empleado = self.dropdown_tipo_empleado.value
        # Actualiza el Dropdown de departamentos basado en el valor seleccionado
        if tipo_empleado == "Confianza" or tipo_empleado=="":
            periodos = obtener_periodo(2)
        elif tipo_empleado == "Sindicato":
            periodos = obtener_periodo(1)
        for i, periodo in enumerate(periodos):
            periodos=periodo[2]
        self.periodos=periodos
        self.update()
        return periodos

    def send_data(self, tipo_d, tipo_e, periodos):
        tipo_e = self.dropdown_tipo_empleado.value
        tipo_d = departamentos

        def manejar_cambio_mes(e):
            # Crear barra de progreso
            # Aquí llamas a la función con los parámetros necesarios
            periodos = obtener_periodos_mes(tipo_e, e.control.value)
            periodos_incompletos = {}
            # Calcular el número total de ciclos
            total_ciclos = len(periodos) * len(tipo_d)
            ciclo_actual = 0
            progress_bar.visible=True
            self.page.update()
            for periodo in periodos:
                for depa in tipo_d:
                    print(f"Verificando periodo: {periodo[2]} para departamento: {depa}")
                    if not excel_add(0, depa, tipo_e, periodo[2]):
                        print("hi")
                        # Si el departamento aún no tiene una lista, se crea
                        if depa not in periodos_incompletos:
                            periodos_incompletos[depa] = []
                        # Agregar el periodo a la lista del departamento
                        periodos_incompletos[depa].append(periodo[2])
                    # Incrementar progreso
                    ciclo_actual += 1
                    progress_bar.value = ciclo_actual / total_ciclos
                    self.page.update()
                    time.sleep(1.5)
            print(periodos_incompletos)
            self.page.update()
            # Mostrar mensaje según el resultado
            if periodos_incompletos:
                contenido = []
                for depa, periodos_faltantes in periodos_incompletos.items():
                    fila_proyecto = ft.Text(f"En el proyecto: {depa}")
                    fila_periodos = ft.Column(
                        controls=[ft.Text(f"{periodo}") for periodo in periodos_faltantes]
                    )
                    contenido.append(ft.Column(controls=[fila_proyecto, ft.Text("En los Periodos:"), fila_periodos]))

                alert = ft.AlertDialog(
                    title=ft.Text("Reportes Generados, Registros Inexistentes o Incompletos:"),
                    content=ft.Column(controls=contenido),scrollable=True
                )
                self.page.dialog = alert  # Asignar el diálogo
                self.page.dialog.open = True  # Abrir el diálogo
                self.page.update()  # Actualizar la página
            else:
                alert = ft.AlertDialog(
                    title=ft.Text("Reporte Generado."),
                    scrollable=True
                )
                self.page.dialog = alert
                self.page.dialog.open = True
                self.page.update()

        # Crear el diálogo de selección de mes
        progress_bar = ft.ProgressBar(width=400, height=20,visible=False)
        meses_dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Seleccione el mes para el reporte"),
            content=ft.Column(controls=[ft.Dropdown(
                bgcolor=ft.colors.GREY_300,
                color=ft.colors.BLACK,
                hint_text="Seleccione el mes...",
                padding=ft.padding.only(left=10, right=10),
                options=[ft.dropdown.Option(i) for i in range(1, 13)],
                on_change=manejar_cambio_mes
            ),progress_bar,])
        )
        # Mostrar el diálogo de selección de mes
        self.page.dialog = meses_dialog
        self.page.dialog.open = True
        self.page.update()
    def bar_icons(self, e):
        # Acción para el icono del botón de inicio (sin uso en este caso)
        pass
def logout(page):
        page.go("/login")   # Redirige a la vista inicial
def encrypt(plain_text):
    str_out = ""
    outx_ = bytearray(len(plain_text))
    idx_ = plain_text.encode('utf-16le')  # Encoding in UTF-16LE (Little Endian) similar to Encoding.Unicode
    Key=Variables().obtener_llave()
    key_idx_ = Key.encode('utf-16le')  # Same encoding for the key
    nbyte = 0
    for n_pos in range(0, len(idx_), 2):
        # Perform XOR between the byte of the plain text and the key
        c = chr(idx_[n_pos] ^ key_idx_[n_pos])
        str_out += c
        outx_[nbyte] = idx_[n_pos] ^ key_idx_[n_pos]
        nbyte += 1
    # Convert the resulting byte array to a base64 string
    return base64.b64encode(outx_).decode('utf-8')
host=socket.gethostname()
departamentos=[]
def obtener_departamentos(usern):
        departamentos.clear()
        db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
        db.conectar()
        resultado=db.ejecutar_consulta("SELECT ID_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE NOMBRE_USUARIO=?",usern)
        id=str(resultado[0][0])
        res=db.ejecutar_consulta("SELECT CLAVE_DEPARTAMENTO FROM HIS_SISTEMAS_DEPUSER WHERE ID_USUARIO=?",id)
        for i in res:
            departamentos.append(str(i[0]))
        db.cerrar()
        return departamentos
def main(page: ft.Page):
    page.add(Text("Welcome"))
    page.go("/login")
    page.window_maximized = True  
    snack = SnackBar(
        Text("Registration successful")
    )
    def verificar_acceso(username, password):
        try:
            global nombre
            nombre=username
            db = ConexionBD(host="148.200.128.15",database="JumapamSistemas",user="andres",password="Andr3s2024")
            db.conectar()
            usuario=username
            password = password
            version = '1.0'
            id_sistema = 12
            host_name = 'DESKTOP'
            cursor=db.cursor
            cursor.execute("{CALL spAccesoSistemas (?, ?, ?, ?, ?, ?, ?)}", 
                    usuario, password, "mac", "ip", version, id_sistema, host_name)
            # Obtener los resultados
            rows = cursor.fetchmany()
            db.commit()
            db.cerrar()
            if "Acceso Correcto|" in str(rows[0]):
                return True
            return False

        except Exception as e:
            logging.error("Error", f"Error al ejecutar el Proceso Almacenado: {e}")
        # Cerrar la conexión
        db.cerrar()
        # Función que maneja el inicio de sesión
    def iniciar_sesion(e, username, password):
        username = username
        if not re.match("^[a-zA-Z0-9_]*$", username):
            alert=ft.AlertDialog(
                title=ft.Text("El nombre contiene caracteres no permitidos."),
                )
            page.open(alert)
            raise ValueError("El nombre contiene caracteres no permitidos.")
        password = encrypt(password)    
        if verificar_acceso(username, password):
            obtener_departamentos(username)
            page.go("/home")
            
        else:
            alert=ft.AlertDialog(
                title=ft.Text("Usuario o Contraseña invalida"),
                )
            page.open(alert)
            page.update()

    def route_change(route):
        username = ft.TextField(
                    width=280,  # Aumenta el ancho del campo para que quede alineado con el contenedor
                    height=40,
                    hint_text="Usuario",
                    border="underline",
                    color="black",
                    prefix_icon=ft.icons.EMAIL,
                )
        password = ft.TextField(
                    width=280,  # Aumenta el ancho del campo para que quede alineado con el contenedor
                    height=40,
                    hint_text="Contraseña",
                    border="underline",
                    color="black",
                    prefix_icon=ft.icons.LOCK,
                    password=True,
                    on_submit=lambda e:iniciar_sesion(e, username.value, password.value)
                )
        conteiner = ft.Container(
        ft.Column([
            # Contenedor de la imagen
            ft.Container(
                ft.Image(
                    src="https://jumapam.gob.mx/images/JPG/jumapam.jpg",  # Cambia a la URL o ruta de tu imagen
                    width=120,
                    height=120,
                    fit=ft.ImageFit.CONTAIN
                ),
                alignment=ft.alignment.center,  # Centra la imagen dentro de su contenedor
                padding=ft.padding.only(20, 20)
            ),
            ft.Container(
                ft.Text(
                    "Iniciar Sesión",
                    color="black",
                    width=320,
                    size=30,
                    text_align="center",
                    weight="w900"
                ),
                padding=ft.padding.only(20, 20),
                alignment=ft.alignment.center
            ),
            ft.Container(
                username,
                padding=ft.padding.only(20, 10),
                alignment=ft.alignment.center  # Centra el campo en el contenedor
            ),
            ft.Container(
                password,
                padding=ft.padding.only(20, 10),
                alignment=ft.alignment.center  # Centra el campo en el contenedor
            ),
            ft.Container(
                ft.ElevatedButton(
                    text="Iniciar",
                    width=280,
                    bgcolor="black",
                    color="white",
                    on_click=lambda e: iniciar_sesion(e, username.value, password.value)
                ),
                padding=ft.padding.only(20, 20),
                alignment=ft.alignment.center  # Centra el botón en el contenedor
            ),
        ],
        alignment=ft.MainAxisAlignment.CENTER,  # Centra verticalmente el contenido en la columna
        horizontal_alignment=ft.CrossAxisAlignment.CENTER  # Centra horizontalmente el contenido en la columna
        ),
        border_radius=20,
        width=360,
        height=500,
        gradient=ft.LinearGradient([
            ft.colors.WHITE,
        ])
    )
        page.views.clear()
        if page.route=="/login":
            page.views.append(
            ft.View(
                    "/login",
                    [
                        conteiner,
                    ],
                    bgcolor = ft.colors.BLUE_900,
                    padding = 0,
                    vertical_alignment = "center",  # Centra verticalmente el contenedor en la pantalla
                    horizontal_alignment = "center",  # Centra horizontalmente el contenedor en la pantalla
                )
        )
        elif page.route=="/home":
            page.views.append(
                ft.View(
                "/home",
                [
                    ft.Text(f"Welcome, {username.value}!!"),
                    AnimatedApp(),
                    ft.ElevatedButton(
                        text="Cerrar sesión",
                        bgcolor=ft.colors.RED_400,
                        color=ft.colors.WHITE,
                        width=140,
                        height=40,
                        on_click=lambda e:logout(page)  # Función para manejar el evento de clic
                    ) 
                ]
                )        
            )
        page.update()

    def view_pop(view):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)
    page.title = 'login'
    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.go(page.route)
    page.update()

ft.app(target=main)
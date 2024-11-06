#Nomina2
import pyodbc
import tkinter as tk
from tkinter import messagebox, ttk  # Agregar ttk para el combobox
from uuid import getnode as get_mac
import base64
from datetime import datetime, timedelta,date
from openpyxl import load_workbook,Workbook
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib import colors,pdfencrypt
from reportlab.lib.pagesizes import letter,landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph,PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl.utils import get_column_letter
from reportlab.lib.units import mm
import socket
varss=[]
host=socket.gethostname()
multireg=False
fechas_seleccionadas = {}
fechas=[]
departamentos=[]
descanso={}
year=datetime.now().year
month=datetime.now().month
day=datetime.now().day
todos={}
dias_festivos=[date(year,1,1),date(year,5,1),date(year,9,16),date(year,12,25)]
dias_semana={0:"LUNES",1:"MARTES",2:"MIERCOLES",3:"JUEVES",4:"VIERNES",5:"SABADO",6:"DOMINGO",}
s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
s.connect(("8.8.8.8", 80))
ip=s.getsockname()[0]
s.close()
def calcular_fechas():
    if ".333" in str(year/6):
        dias_festivos.append(date(year,10,1))
    d = date(year, 2, 1)
    offset = 0-d.weekday() #weekday = 0 means monday
    if offset < 0:
        offset+=7
    dias_festivos.append(date(year,2,1+offset))
    d=date(year,3,14)
    offset=0+-d.weekday()
    if offset < 0:
        offset+=7
    dias_festivos.append(date(year,3,14+offset))
    d=date(year,11,14)
    offset=0+-d.weekday()
    if offset < 0:
        offset+=7
    dias_festivos.append(date(year,11,14+offset))
    return d+timedelta(offset)
calcular_fechas()
# Define your encryption key here (equivalente a KEY_ENCRYPT_DECRYPT en el código original)
KEY_ENCRYPT_DECRYPT = "r3c6rs0sm4t3r14l3sj6m4p4mm4z4tl4ns1n4l04l4p13ld3lm4rr3c6rs0sm4t3r14l3sj6m4p4mm4z4tl4ns1n4l04l4p13ld3lm4r"
mac = get_mac()
user=''
f=3
def desbut(emp_id):
    state = str(todos[emp_id]['boton'].button['state'])
    if state == 'disabled':
        todos[emp_id]['boton'].config(state="enabled")
    else:
        todos[emp_id]['boton'].config(state="disabled")
def callback():
    global buttonClicked
    buttonClicked = not buttonClicked 
buttonClicked  = False # Bfore first click
def sortdias(dias):
    a={}
    for i,dia in enumerate(dias):
        if dia ==5 or dia ==6:
            a.update({(dias_semana[dia],i):tk.IntVar(value=1)})
        else:
            a.update({(dias_semana[dia],i):tk.IntVar(value=0)})
    return a 
def encrypt(plain_text):
    str_out = ""
    outx_ = bytearray(len(plain_text))
    idx_ = plain_text.encode('utf-16le')  # Encoding in UTF-16LE (Little Endian) similar to Encoding.Unicode
    key_idx_ = KEY_ENCRYPT_DECRYPT.encode('utf-16le')  # Same encoding for the key
    nbyte = 0
    for n_pos in range(0, len(idx_), 2):
        # Perform XOR between the byte of the plain text and the key
        c = chr(idx_[n_pos] ^ key_idx_[n_pos])
        str_out += c
        outx_[nbyte] = idx_[n_pos] ^ key_idx_[n_pos]
        nbyte += 1
    # Convert the resulting byte array to a base64 string
    return base64.b64encode(outx_).decode('utf-8')
# Función para conectar a la base de datos de usuarios
def conectar_bd_usuarios():
    conexion = pyodbc.connect(
        'DRIVER={SQL Server};'
        f'SERVER={host}\SQLEXPRESS;'
        'DATABASE=JumapamSistemas;'  # Base de datos de usuarios
        'Trusted_Connection=yes;'
    )
    cursor = conexion.cursor()
    return conexion, cursor
def conectar_bd_periodos():
    conexion = pyodbc.connect(
        'DRIVER={SQL Server};'
        f'SERVER={host}\SQLEXPRESS;'
        'DATABASE=BdTrabajadTemporal;'  # Base de datos de usuarios
        'Trusted_Connection=yes;'
    )
    cursor = conexion.cursor()
    return conexion, cursor
def conectar_bd_Jumapam():
    conexion = pyodbc.connect(
        'DRIVER={SQL Server};'
        f'SERVER={host}\SQLEXPRESS;'
        'DATABASE=Jumapam;' 
        'Trusted_Connection=yes;'
    )
    cursor = conexion.cursor()
    return conexion, cursor
# Función para conectar a la base de datos de datos
def conectar_bd_datos():
    conexion = pyodbc.connect(
        'DRIVER={SQL Server};'
        f'SERVER={host}\SQLEXPRESS;'
        'DATABASE=datos;'  # Base de datos donde se guardan los datos
        'Trusted_Connection=yes;'
    )
    cursor = conexion.cursor()
    return conexion, cursor
def obtener_quincenas():
    a=[]
    if day <= 15:
        # Primera quincena (del 1 al 15)
        inicio = date(year, month, 1)
        fin = date(year, month, 15)
        for i in range(15):
            a.append((inicio+timedelta(i)).weekday())
    else:
        # Segunda quincena (del 16 al último día del mes)
        inicio = date(year, month, 16)
        ultimo_dia = (inicio.replace(month=month % 12 + 1, day=1) - timedelta(days=1)).day
        fin = date(year, month, ultimo_dia)
        for i in range(ultimo_dia-15):
            a.append((inicio+timedelta(i)).weekday())    
    return a
def almacenar_fecha(dia_semana, var_checkbox, dia_texto, nomina):
    # Obtener el día de la semana actual
    hoy = datetime.now()
    # Calcular la fecha del día seleccionado
    if nomina == "2" and day <=15:
        fecha = hoy - timedelta(days=day) + timedelta(days=dia_semana+1)
    elif nomina=="2" and day>15:
        fecha=datetime(year,month,16)+timedelta(days=dia_semana)
    else:
        fecha = hoy - timedelta(days=hoy.weekday()) + timedelta(days=dia_semana+0)
    # Formatear la fecha como DD/MM/AAAA
    fecha_formateada = fecha.strftime('%d/%m/%Y')
    if var_checkbox.get()==1:
        # Si el checkbox está marcado, agregamos la fecha
        fechas_seleccionadas[dia_texto,dia_semana] = fecha_formateada
    else:
        # Si se desmarca, quitamos la fecha de la lista
        if (dia_texto,dia_semana) in fechas_seleccionadas:
            del fechas_seleccionadas[dia_texto,dia_semana]
def excel():
    hoy = datetime.now()
    fechas = []
    # Calcular la fecha del día seleccionado
    for i in range(7):
        fecha = hoy - timedelta(days=hoy.weekday()) + timedelta(days=i)
        fecha_formateada = fecha.strftime('%d/%m/%Y')
        fechas.append(fecha_formateada)
    # Crear un nuevo libro de Excel o cargar uno existente
    ruta_excel = "C:\\Users\\usuario\\Downloads\\Nomina\\Formato.xlsx"
    # Verificar si el archivo existe y cargarlo, de lo contrario, crear uno nuevo
    try:
        workbook = load_workbook(ruta_excel)
    except FileNotFoundError:
        workbook = Workbook()
    # Obtener la hoja de trabajo (el índice empieza en cero, o usa el nombre de la hoja)
    sheet = workbook.active  # O usa workbook["NombreHoja"] para acceder a una hoja específica
    # Asignar fechas a las celdas correspondientes
    for i in range(7):
        column_letter = get_column_letter(5 + i)  # E es la columna 5
        sheet[f"{column_letter}3"] = fechas[i]
    # Guardar el archivo de Excel actualizado
    workbook.save(ruta_excel)
    # Configuración de la página
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.printGridlines = True
    workbook.close()
def verificar_asistencias(codigoEmpleado,periodo):
    try:
        # Conectar a la base de datos
        conexion, cursor = conectar_bd_datos()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        cursor.execute("SELECT Codigo_Empleado,Dia_Asistencia,Horas_Extra,Turnos_Extras,Descansos_Trabajados FROM Datos1 WHERE Codigo_Empleado=? AND Periodo=?",codigoEmpleado,periodo)
        resultado = cursor.fetchall()
        asistencia_modificada = []
        for fila in resultado:
            print(fila[1])
            if fila[1] != "0": 
                dia_asistencia = 1
            else:
                dia_asistencia = 0
            asistencia_modificada.append((fila[0], dia_asistencia, fila[2], fila[3], fila[4]))
        conexion.close()
        return asistencia_modificada if asistencia_modificada else None  # Asegurarse de devolver None si no hay datos
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo añadir o actualizar el dato: {e}")
def excel_add(id,depar,tipo,periodo):
    # Cargar el archivo de Excel
    ruta_excel = "C:\\Users\\usuario\\Downloads\\Nomina\\Formato.xlsx"
    workbook = load_workbook(ruta_excel)
    # Obtener la hoja de trabajo
    sheet = workbook.active 
    x=0
    empleados = obtener_empleados(depar,tipo)
    nombres = []
    ape=[]
    ape2=[]
    # Extraer el nombre completo del empleado
    for empleado in empleados:
        nombre_completo = f"{empleado[2]}"
        nombres.append(nombre_completo)
        ap=f"{empleado[3]}"
        ape.append(ap)
        ap2=f"{empleado[4]}"
        ape2.append(ap2)
        x+=1
    for i in range(4, x + 4):
        Dato = verificar_asistencias(i - 3,periodo)
        if sheet["A" + str(i)].value is None and Dato and len(Dato) > 0:
            sheet["A" + str(i)] = str(Dato[0][0])  # Código de empleado
        sheet["B" + str(i)] = str(nombres[i - 4])  # Nombre completo del empleado
        sheet["C" + str(i)] = str(ape[i - 4])  # Nombre completo del empleado
        sheet["D" + str(i)] = str(ape2[i - 4])  # Nombre completo del empleado
        # Solo asignar si hay datos en Dato[0] y Dato[0][1]
        sheet["E" + str(i)] = 1 if Dato[0][1] > 0 else 0  # Dias_Asistencia (0 o 1)
        sheet["L" + str(i)] = str(Dato[0][2])  # Horas_Extra
        sheet["M" + str(i)] = str(Dato[0][3])  # Turnos_Extras
        sheet["N" + str(i)] = str(Dato[0][4])  # Descansos_Trabajados
        # Asegúrate de verificar la longitud de Dato[1] y demás
        sheet["F" + str(i)] = str(Dato[1][1])  # Datos de asistencia
        sheet["O" + str(i)] = str(Dato[1][2])  # Más datos
        sheet["P" + str(i)] = str(Dato[1][3])  # Más datos
        sheet["Q" + str(i)] = str(Dato[1][4])  # Más datos
        sheet["G" + str(i)] = str(Dato[2][1])  # Datos adicionales
        sheet["R" + str(i)] = str(Dato[2][2])  # Más datos
        sheet["S" + str(i)] = str(Dato[2][3])  # Más datos
        sheet["T" + str(i)] = str(Dato[2][4])  # Más datos
        sheet["H" + str(i)] = str(Dato[3][1])  # Más datos
        sheet["U" + str(i)] = str(Dato[3][2])  # Más datos
        sheet["V" + str(i)] = str(Dato[3][3])  # Más datos
        sheet["W" + str(i)] = str(Dato[3][4])  # Más datos
        sheet["I" + str(i)] = str(Dato[4][1])  # Más datos
        sheet["X" + str(i)] = str(Dato[4][2])  # Más datos
        sheet["Y" + str(i)] = str(Dato[4][3])  # Más datos
        sheet["Z" + str(i)] = str(Dato[4][4])  # Más datos
        sheet["J" + str(i)] = str(Dato[5][1])  # Más datos
        sheet["AA" + str(i)] = str(Dato[5][2])  # Más datos
        sheet["AB" + str(i)] = str(Dato[5][3])  # Más datos
        sheet["AC" + str(i)] = str(Dato[5][4])  # Más datos
        sheet["K" + str(i)] = str(Dato[6][1])  # Más datos
        sheet["AD" + str(i)] = str(Dato[6][2])  # Más datos
        sheet["AE" + str(i)] = str(Dato[6][3])  # Más datos
        sheet["AF" + str(i)] = str(Dato[6][4])  # Más datos
        for j in range(1, 8):  # Ajusta según la cantidad de Dato
                col_letter = chr(70 + j)  # Calcula la letra de la columna
                if sheet[col_letter + str(i)].value is None and len(Dato) > j:
                    sheet[col_letter + str(i)] = str(Dato[j])  # Asigna el valor
    # Ajustar el ancho de las columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            if cell.value:  # Solo verificar celdas que no están vacías
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)  # Ajustar ligeramente
        sheet.column_dimensions[col_letter].width = adjusted_width
    # Guardar el archivo de Excel actualizado
    workbook.save("C:\\Users\\usuario\\Downloads\\Nomina\\Formatollenado.xlsx")
    periodo=periodo.replace(" ", "_")
    periodo=periodo.replace("/", "-")
    # Convertir a PDF usando pandas y matplotlib
    ruta_excel = "C:\\Users\\usuario\\Downloads\\Nomina\\Formatollenado.xlsx"
    pdf_output = "C:\\Users\\usuario\\Downloads\\Nomina\\output\\Reporte_"+periodo
    excel_to_pdf(ruta_excel, pdf_output)   
def excel_to_pdf(excel_file, pdf_file):
    # Leer el archivo de Excel
    df = pd.read_excel(excel_file)
    df = df.fillna("")  # Reemplazar valores nulos con cadenas vacías
    # Crear un documento PDF con orientación horizontal y márgenes pequeños
    pdf = SimpleDocTemplate(
        pdf_file+".pdf", 
        pagesize=landscape(letter), 
        leftMargin=10, 
        rightMargin=10, 
        topMargin=10, 
        bottomMargin=10, encrypt=pdfencrypt.StandardEncryption("pass", canPrint=0,canModify=0,canCopy=0,canAnnotate=0)
    )
    # Estilo para los párrafos (ajuste de texto)
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    # Dividir los datos en dos partes:
    # 1. Columnas de A a K
    df_A_K = df.iloc[:, :11]  # Columnas A a K son las primeras 11 (indexado desde 0)
    # 2. Unir columnas de A a D y L en adelante
    df_A_D_L_onward = pd.concat([df.iloc[:, :1], df.iloc[:, 11:]], axis=1)  # Unir columnas A-D y L en adelante
    # Función para crear tablas a partir de dataframes y añadirlas al documento
    def create_table(data, num_cols):
        # Convertir los datos a lista de listas y luego a Paragraphs para ajuste de texto
        table_data = []
        for row in data.astype(str).values.tolist():
            new_row = [Paragraph(cell, style_normal) for cell in row]
            table_data.append(new_row)
        # Calcular los tamaños de las columnas y filas
        page_width, page_height = landscape(letter)
        total_width = page_width - 20  # Restar márgenes
        col_width = total_width / num_cols  # Ancho de cada columna
        # Crear tabla con las columnas ajustadas
        table = Table(table_data, colWidths=[col_width] * num_cols)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Fondo del encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),  # Reducir el tamaño de la letra
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alinear verticalmente
        ])
        table.setStyle(style)
        return table
    # Construir el PDF
    elements = []
    # Primera página: columnas A a K
    elements.append(create_table(df_A_K, 11))  # 11 columnas de A a K
    # Saltar a la siguiente página
    elements.append(PageBreak())
    # Segunda página: combinación de columnas A a D y L en adelante
    elements.append(create_table(df_A_D_L_onward, len(df_A_D_L_onward.columns)))
    # Generar el PDF
    pdf.build(elements)
# Función para verificar si el usuario existe en la tabla de usuarios
def verificar_acceso(username, password):
    try:
        conexion, cursor = conectar_bd_usuarios()
        usuario = username
        password = password
        version = '1.0'
        id_sistema = 12
        host_name = 'DESKTOP'
        cursor.execute("{CALL spAccesoSistemas (?, ?, ?, ?, ?, ?, ?)}", 
                   usuario, password, mac, ip, version, id_sistema, host_name)
        # Obtener los resultados
        rows = cursor.fetchmany()
        conexion.commit()
        if "Acceso Correcto|" in str(rows[0]):
            return True
        return False

    except Exception as e:
        messagebox.showerror("Error", f"Error al ejecutar el Proceso Almacenado: {e}")
    # Cerrar la conexión
    if conexion:
        conexion.close()
# Función que maneja el inicio de sesión
def iniciar_sesion():
    username = entry_username.get()
    password = encrypt(entry_password.get())
    if verificar_acceso(username, password):
        messagebox.showinfo("Éxito", "Acceso concedido. Usuario y contraseña correctos.")
        root.destroy()  # Cierra la ventana de login
        abrir_ventana_principal(username, 'Sindicato')  # Abre la ventana para añadir datos
    else:
        messagebox.showerror("Error", "Usuario o contraseña incorrectos.")
# Función para abrir la ventana principal donde añadir datos
def obtener_empleados(depar,tipo):
    if tipo== "Sindicato":
        tipoc=2
    else:
        tipoc=1
    conexion, cursor = conectar_bd_periodos()
    print(depar,tipoc)
    cursor.execute("SELECT TRABAJAD.CLAVE_TRABAJADOR,CLAVE_TIPO_NOMINA, NOMBRE, PATERNO, MATERNO, DESCANSO1, DESCANSO2,CLAVE_DEPARTAMENTO FROM TRABAJAD INNER JOIN TRAHISDE ON TRAHISDE.CLAVE_TRABAJADOR=TRABAJAD.CLAVE_TRABAJADOR WHERE FECHA_F='2100-12-31' AND CLAVE_DEPARTAMENTO=? AND CLAVE_TIPO_NOMINA=?",depar,tipoc)
    empleados = cursor.fetchall()
    conexion.close()
    return empleados
def obtener_periodo(x):
    hoy=date.today()
    if x==1:
        conexion, cursor = conectar_bd_periodos()
        cursor.execute("SELECT CLAVE_PERIODO, CLAVE_TIPO_NOMINA, DESCRIPCION FROM PERIODO WHERE CLAVE_TIPO_NOMINA=1 AND ? BETWEEN FECHA_I AND FECHA_F",str(hoy))
        periodo = cursor.fetchall()
        conexion.close()
        return periodo
    else:
        conexion, cursor = conectar_bd_periodos()
        cursor.execute("SELECT CLAVE_PERIODO, CLAVE_TIPO_NOMINA, DESCRIPCION FROM PERIODO WHERE CLAVE_TIPO_NOMINA=2 AND ? BETWEEN FECHA_I AND FECHA_F",str(hoy))
        periodo = cursor.fetchall()
        conexion.close()
        return periodo
def obtener_departamentos(user):
    conexion, cursor = conectar_bd_usuarios()
    cursor.execute("SELECT ID_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE NOMBRE_USUARIO=?",user)
    r=cursor.fetchone()
    id= str(r[0])
    cursor.execute("SELECT CLAVE_DEPARTAMENTO FROM HIS_SISTEMAS_DEPUSER WHERE ID_USUARIO=?",id)
    for i in cursor:
        departamentos.append(str(i[0]))
    conexion.close()
    return departamentos
def obtener_ALL_departamentos():
    conexion, cursor = conectar_bd_periodos()
    cursor.execute("SELECT CLAVE_DEPARTAMENTO FROM DEPARTAM")
    depars=[]
    for i in cursor:
        depars.append(str(i[0]))
    conexion.close()
    return depars
# Función para crear la interfaz y mostrar los empleados con checkboxes y comentario
def verificar_Permisos(usuario):
    try:
        conexion, cursor = conectar_bd_usuarios()
        cursor.execute("SELECT ID_USUARIO FROM MAE_SISTEMAS_USUARIOS WHERE NOMBRE_USUARIO = ?",usuario)
        id=cursor.fetchall()
        conexion.close()
        conexion, cursor = conectar_bd_usuarios()
        cursor.execute("{CALL spAccesoSistemasPermisos (?)}",id[0])
        # Obtener los resultados
        rows = cursor.fetchall()
        i=0
        for row in rows:
            i=i+1
            if i>12:
                if row[11]:
                    return True
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error al ejecutar el Proceso alamcenado: {e}")
    # Cerrar la conexión
    if conexion:
        conexion.close()
def GET_USER():
    try:
        conexion, cursor = conectar_bd_usuarios()
        cursor.execute("SELECT ID_USUARIO FROM HIS_SISTEMAS_PERMISOS WHERE ID_SISTEMA = 12")
        id=cursor.fetchall()
        iD=[]
        for ids in id:
            print(ids)
            iD.append(str(ids[0]))
        conexion.close()
        return iD
    except Exception as e:
        messagebox.showerror("Error", f"Error al ejecutar el Proceso alamcenado: {e}")
    # Cerrar la conexión
    if conexion:
        conexion.close()
def UPDATE_USER(id,clave):
    try:
        conexion, cursor = conectar_bd_usuarios()
        for clavedep in clave:
            print(clavedep)
            cursor.execute("SELECT ID_USuARIO FROM HIS_SISTEMAS_DEPUSER WHERE ID_USUARIO = ? AND CLAVE_DEPARTAMENTO=?",id,clavedep)
            if cursor.fetchone() is None:
                cursor.execute("INSERT INTO HIS_SISTEMAS_DEPUSER VALUES(?,?)",id,clavedep)
                cursor.commit()
            else:
                print(cursor.fetchone())
        cursor.close()
    except Exception as e:
        messagebox.showerror("Error", f"Error al ejecutar el Proceso alamcenado: {e}")
    # Cerrar la conexión
    if conexion:
        conexion.close()
def checkdep(id, clave,vars):
    try:
        conexion, cursor = conectar_bd_usuarios()
        for i,(clavedep,var) in enumerate(clave):
            cursor.execute("SELECT ID_USuARIO FROM HIS_SISTEMAS_DEPUSER WHERE ID_USUARIO = ? AND CLAVE_DEPARTAMENTO=?",id,clavedep)
            if cursor.fetchone() is None:
                print("BYE")
                vars[i]=tk.IntVar(value=0)
            else:
                print("HI")
                vars[i]=tk.IntVar(value=1)
        cursor.close()
    except Exception as e:
        messagebox.showerror("Error", f"Error al ejecutar: {e}")
    # Cerrar la conexión
    if conexion:
        conexion.close()
def abrir_ventana_principal(username, tipo_opcion):
    departamentos = obtener_departamentos(username)
    print(departamentos)
    empleados = obtener_empleados(departamentos[0],tipo_opcion)
    permiso = verificar_Permisos(username)
    # Crear la ventana
    ventana_empleados = tk.Tk()
    ventana_empleados.title("Empleados - Registro prenomina")
    ventana_empleados.state('zoomed')
    # Botón de cerrar sesión
    button_logout = tk.Button(ventana_empleados, text="Cerrar sesión", command=lambda: cerrar_sesion(ventana_empleados))
    button_logout.grid(row=2, column=4, sticky="w")
    # Botón de agregar
    button_user = tk.Button(ventana_empleados, text="Agregar Usuario", command=lambda: Agregar_usuario(ventana_empleados))
    button_user.grid(row=2, column=5, sticky="w")
    # Combobox para seleccionar "Confianza" o "Sindicato"
    label_opcion = tk.Label(ventana_empleados, text="Selecciona la opción:")
    label_opcion.grid(row=2, column=0, padx=5, pady=10, sticky="w")
    combobox = ttk.Combobox(ventana_empleados, values=["Sindicato", "Confianza"])
    combobox.set(tipo_opcion)  # Establecer la opción por defecto
    combobox.grid(row=2, column=1, padx=5, pady=10, sticky="w")
    label_opcion = tk.Label(ventana_empleados, text="Selecciona Departamento:")
    label_opcion.grid(row=2, column=2, padx=5, pady=10, sticky="w")
    combobox2 = ttk.Combobox(ventana_empleados, values=departamentos)
    combobox2.set(departamentos[0])  # Establecer la opción por defecto
    combobox2.grid(row=2, column=3, padx=5, pady=10, sticky="w")
    combobox2['state'] = "readonly"
    combobox['state'] = "readonly"
    # Configurar el comportamiento de las filas y columnas
    ventana_empleados.grid_rowconfigure(1, weight=1)
    ventana_empleados.grid_columnconfigure(0, weight=1)
    # Crear el Canvas y el Frame dinámico
    canvas = tk.Canvas(ventana_empleados)
    canvas.grid(row=1, column=0, sticky="nsew")
    frame_dinamico = tk.Frame(canvas)
    frame_dinamico.grid(row=0, column=0, sticky="nsew")
    # Asociar el Frame dinámico con el Canvas
    canvas.create_window((0, 0), window=frame_dinamico, anchor=tk.NW)
    # Inicialización de la tabla
    actualizar_contenido(ventana_empleados, frame_dinamico, empleados, tipo_opcion, permiso, departamentos[0])
    # Vincular la combobox
    combobox.bind("<<ComboboxSelected>>", lambda event: actualizar_contenido(ventana_empleados, frame_dinamico, obtener_empleados(combobox2.get(),combobox.get()), combobox.get(), permiso, combobox2.get()))
    combobox2.bind("<<ComboboxSelected>>", lambda event: actualizar_contenido(ventana_empleados, frame_dinamico, obtener_empleados(combobox2.get(),combobox.get()), combobox.get(), permiso, combobox2.get()))
    # Ajustar el tamaño del canvas cuando se redimensiona
    def ajustar_canvas(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    ventana_empleados.bind("<Configure>", ajustar_canvas)
    ventana_empleados.mainloop()
# Función para agregar datos a la base de datos
def agregar_dato(dias, comentario, periodo, aprovacion, codigoEmpleado, HE, DF, TE, DT,descanso,nomina):
    global buttonClicked
    if nomina=="2":
        for i, ((dia,index), var) in enumerate(dias.items()):
            almacenar_fecha(i, var, dia, nomina)  # Guarda las fechas seleccionadas
    else:
        for i, (dia, var) in enumerate(dias.items()):
            almacenar_fecha(i, var, dia, nomina)  # Guarda las fechas seleccionadas
    try:
        # Conectar a la base de datos
        conexion, cursor = conectar_bd_datos()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        cursor.execute("SELECT COUNT(*) FROM Datos1 WHERE Codigo_Empleado = ? AND periodo= ?", (codigoEmpleado,periodo))
        resultado = cursor.fetchone()
        DF = "0"
        # Iterar sobre los días festivos
        for dia_festivo in dias_festivos:
            if date.today() == dia_festivo:
                DF = "1"
                break  
        if resultado[0] == 0:
            # Si no existe, insertar un nuevo registro
            if nomina=="2":
                for i, ((dia,index), var) in enumerate(dias.items()):
                    if dia in str(descanso) and var.get():
                        DT[i]="1"
                    cursor.execute(
                        """
                        INSERT INTO Datos1 
                        (Codigo_Empleado, TipoCobro, Dia_Semana, Dia_Asistencia, Horas_Extra, Dias_Festivos, Turnos_Extras, Descansos_Trabajados, Periodo, Aprobacion) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            codigoEmpleado,
                            nomina,           # Tipo de cobro
                            dia,                   # Día de la semana
                            str(fechas_seleccionadas.get((dia,i), 0)),  # Asistencia del día
                            HE.get(i, "0"),         # Horas extra del día i
                            DF,                    # Días festivos (DF) es fijo
                            TE.get(i, "0"),         # Turnos extras del día i
                            DT.get(i, "0"),         # Descansos trabajados del día i
                            periodo,               # Periodo actual
                            aprovacion             # Aprobación
                        )
                    )
            else:
                for i, (dia, var) in enumerate(dias.items()):
                    if dia in str(descanso) and var.get():
                        DT[i]="1"
                    cursor.execute(
                        """
                        INSERT INTO Datos1 
                        (Codigo_Empleado, TipoCobro, Dia_Semana, Dia_Asistencia, Horas_Extra, Dias_Festivos, Turnos_Extras, Descansos_Trabajados, Periodo, Aprobacion) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            codigoEmpleado,
                            nomina,           # Tipo de cobro
                            dia,                   # Día de la semana
                            str(fechas_seleccionadas.get((dia,i), 0)),  # Asistencia del día
                            HE.get(i, "0"),         # Horas extra del día i
                            DF,                    # Días festivos (DF) es fijo
                            TE.get(i, "0"),         # Turnos extras del día i
                            DT.get(i, "0"),         # Descansos trabajados del día i
                            periodo,               # Periodo actual
                            aprovacion             # Aprobación
                        )
                    )
            conexion.commit()
            if multireg==True and buttonClicked==True:
                messagebox.showinfo("Éxito", "Multiples Datos añadidos correctamente.")
                buttonClicked=False
            elif buttonClicked==False:
                messagebox.showinfo("Éxito", "Datos añadidos correctamente.")
        else:
            # Si ya existe, realizar un update
            if nomina=="2":
                for i, ((dia,index), var) in enumerate(dias.items()):
                    if dia in str(descanso) and var.get():
                        DT[i]="1"
                    cursor.execute(
                        """
                        UPDATE Datos1 SET 
                        Dia_Asistencia = ?,
                        Horas_Extra = ?, Dias_Festivos = ?, Turnos_Extras = ?, Descansos_Trabajados = ?, Aprobacion = ?
                        WHERE Codigo_Empleado = ? AND Dia_Semana = ? AND Periodo = ?
                        """,
                        (
                            str(fechas_seleccionadas.get((dia,i), 0)),  # Asistencia del día
                            HE.get(i, ""),         # Horas extra del día i
                            DF,                    # Días festivos (DF) es fijo
                            TE.get(i, ""),         # Turnos extras del día i
                            DT.get(i, ""),         # Descansos trabajados del día i
                            aprovacion,            # Aprobación
                            codigoEmpleado,        # Código del empleado
                            dia,                   # Día de la semana
                            periodo                # Periodo actual
                        )
                    )
            else:
                for i, (dia, var) in enumerate(dias.items()):
                    if dia in str(descanso) and var.get():
                        DT[i]="1"
                    cursor.execute(
                        """
                        UPDATE Datos1 SET 
                        Dia_Asistencia = ?,
                        Horas_Extra = ?, Dias_Festivos = ?, Turnos_Extras = ?, Descansos_Trabajados = ?, Aprobacion = ?
                        WHERE Codigo_Empleado = ? AND Dia_Semana = ? AND Periodo = ?
                        """,
                        (
                            str(fechas_seleccionadas.get((dia,i), 0)),  # Asistencia del día
                            HE.get(i, ""),         # Horas extra del día i
                            DF,                    # Días festivos (DF) es fijo
                            TE.get(i, ""),         # Turnos extras del día i
                            DT.get(i, ""),         # Descansos trabajados del día i
                            aprovacion,            # Aprobación
                            codigoEmpleado,        # Código del empleado
                            dia,                   # Día de la semana
                            periodo                # Periodo actual
                        )
                    )
            conexion.commit()
            if multireg==True and buttonClicked==True:
                messagebox.showinfo("Éxito", "Multiples Datos añadidos correctamente.")
                buttonClicked=False
            elif buttonClicked==False:
                messagebox.showinfo("Éxito", "Datos actualizados correctamente.")
        conexion.close()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo añadir o actualizar el dato: {e}")
def checar_aprovacion(codigoEmpleado,periodo):
    try:
        # Conectar a la base de datos
        conexion, cursor = conectar_bd_datos()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        cursor.execute("SELECT Aprobacion FROM Datos1 WHERE Codigo_Empleado = ? And Periodo= ?",codigoEmpleado,periodo)
        resultado = cursor.fetchone()
        if "0" in str(resultado) or resultado == None:
            return False
        conexion.close()    
        return True
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo añadir o actualizar el dato: {e}")
def check_dias(dia,codigoempleado,periodo):
    try:
        # Conectar a la base de datos
        conexion, cursor = conectar_bd_datos()
        # Verificar si ya existe un registro con el mismo codigoEmpleado
        cursor.execute("SELECT Dia_Asistencia FROM Datos1 Where Codigo_Empleado=? AND Dia_Semana=? AND Periodo=?",codigoempleado,dia,periodo)
        resultado = cursor.fetchone()
        if resultado == None:
            conexion.close()
            if dia == "SABADO" or dia=="DOMINGO":
                return 0
            return 1
        elif "/" not in str(resultado):
            conexion.close()
            return 0
        conexion.close()    
        return 1
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo añadir o actualizar el dato: {e}")
# Función para cambiar la opción seleccionada en el ComboBox
def actualizar_contenido(ventana, frame_dinamico, empleados, tipo_opcion, fv, depar):
    # Elimina todo el contenido actual del frame
    for widget in frame_dinamico.winfo_children():
        widget.destroy()
    # Etiquetas de encabezado de la tabla
    if tipo_opcion == "Confianza":
        periodo = obtener_periodo(1)
    else:
        periodo = obtener_periodo(2)
    for i, periodo in enumerate(periodo):
        label1 = tk.Label(frame_dinamico, text=periodo[2], font=('Arial', 10, 'bold'))
        label1.grid(row=0, column=1, padx=5, pady=5)
    entries_HE = {}  # Diccionario para almacenar los widgets Entry de HE
    entries_DT = {}  # Diccionario para almacenar los widgets Entry de DT
    entries_TE = {}  # Diccionario para almacenar los widgets Entry de TE
    varaprob = {}
    # Encabezados fijos y desplazables
    headers_fijos = ['ID', 'Nombre', 'Proyecto']
    headers_scrollables = ['Días de la Semana', 'Comentario', 'Aprobar', 'Acción']
    headers_dias = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    # Frame para las columnas fijas
    frame_fijo = tk.Frame(frame_dinamico)
    frame_fijo.grid(row=2, column=0, sticky="n")
    # Botón de generar reporte
    button_excel = tk.Button(frame_dinamico, text="Generar reporte", command=lambda: excel_add(empleado_id,depar,tipo_opcion,periodo[2]))
    button_excel.grid(row=4, column=0, sticky="w")
    # Crear el encabezado de las columnas fijas
    for i, header in enumerate(headers_fijos):
        label = tk.Label(frame_fijo, text=header, font=('Arial', 10, 'bold'))
        label.grid(row=0, column=i, padx=5, pady=10)
    # Canvas para las columnas desplazables
    canvas_scrollable = tk.Canvas(frame_dinamico)
    canvas_scrollable.grid(row=2, column=1, sticky="nsew")  # Ajustado para expandirse
    # Configurar grid para que se expanda
    frame_dinamico.grid_columnconfigure(1, weight=1)  # Columna 1 se expande
    frame_dinamico.grid_rowconfigure(2, weight=1)     # Fila 2 se expande
    # Scrollbar horizontal
    scrollbar_horizontal = tk.Scrollbar(frame_dinamico, orient="horizontal", command=canvas_scrollable.xview)
    scrollbar_horizontal.grid(row=3, column=1, sticky="ew")
    # Scrollbar vertical
    scrollbar_vertical = tk.Scrollbar(frame_dinamico, orient="vertical", command=canvas_scrollable.yview)
    scrollbar_vertical.grid(row=2, column=2, sticky="ns")
    # Frame dentro del canvas desplazable
    frame_scrollable = tk.Frame(canvas_scrollable)
    canvas_scrollable.create_window((0, 0), window=frame_scrollable, anchor="nw")
    canvas_scrollable.configure(xscrollcommand=scrollbar_horizontal.set, yscrollcommand=scrollbar_vertical.set)
    # Ajustar la región de scroll del canvas
    def ajustar_scroll(event):
        canvas_scrollable.configure(scrollregion=canvas_scrollable.bbox("all"))
    frame_scrollable.bind("<Configure>", ajustar_scroll)
    # Crear el encabezado de las columnas desplazables
    for i, header in enumerate(headers_scrollables):
        label = tk.Label(frame_scrollable, text=header, font=('Arial', 10, 'bold'))
        if header == "Días de la Semana":
            label.grid(row=0, column=i, padx=5, pady=10, columnspan=7)
        else:
            label.grid(row=0, column=i+6, padx=5, pady=10)
    frame_fijo2 = tk.Frame(frame_scrollable)
    frame_fijo2.grid(row=1, column=0, sticky="nsew")
    diasq=obtener_quincenas()
    if tipo_opcion == "Confianza":
        for i, header in enumerate(headers_dias):
            label = tk.Label(frame_scrollable, text=header, font=('Arial', 10, 'bold'))
            label.grid(row=2, column=i, padx=5)
    else:
        for i, header in enumerate(diasq):
            label = tk.Label(frame_scrollable, text=dias_semana[header], font=('Arial', 10, 'bold'))
            label.grid(row=2, column=i, padx=5)
    # Crear las filas de datos
    tk.Label(frame_fijo, text="----------------------------------------").grid(row=1, column=0, columnspan=4, padx=5, pady=3)
    dias_seleccionados={}
    dias_seleccionados.clear()
    todos.clear()
    for index, empleado in enumerate(empleados, start=3):
        if empleado[7] == depar:
            empleado_id = empleado[0]
            nombre = empleado[2] + " " + empleado[3]
            proyecto = empleado[7]
            Nomina = empleado[1]
            descanso = {empleado[0], empleado[5]}
            # Mostrar datos en columnas fijas
            tk.Label(frame_fijo, text=empleado_id).grid(row=index, column=0, padx=5, pady=14)
            tk.Label(frame_fijo, text=nombre, font=('Arial', 8, 'bold')).grid(row=index, column=1, padx=5, pady=14)
            tk.Label(frame_fijo, text=proyecto).grid(row=index, column=2, padx=5, pady=14)
            dias_seleccionados={}
            # Crear checkbox para los días de la semana
            if tipo_opcion == "Confianza":
                dias_seleccionados = {
                    'LUNES': tk.IntVar(value=1),
                    'MARTES': tk.IntVar(value=1),
                    'MIERCOLES': tk.IntVar(value=1),
                    'JUEVES': tk.IntVar(value=1),
                    'VIERNES': tk.IntVar(value=1),
                    'SABADO': tk.IntVar(),
                    'DOMINGO': tk.IntVar()
                }
                for i, (dia, var) in enumerate(dias_seleccionados.items()):
                    a = tk.IntVar(value=check_dias(dia, empleado_id,periodo[2]))
                    dias_seleccionados[dia] = a
            else:
                dias_seleccionados = sortdias(diasq)
                for i, (dia, var) in enumerate(dias_seleccionados.items()):
                    a = tk.IntVar(value=check_dias(dia[0], empleado_id,periodo[2]))
                    dias_seleccionados[dia] = a
            frames = []
            for i, (dia, var) in enumerate(dias_seleccionados.items()):
                frame = tk.Frame(frame_scrollable, borderwidth=1)
                frame.grid(row=index, column=i)
                frames.append(frame)
            for i, (dia, var) in enumerate(dias_seleccionados.items()):
                checkbox = tk.Checkbutton(frames[i], variable=var, command=lambda i=i, var=var, dia=dia, emp_id=empleado_id: almacenar_fecha(i, var, dia, Nomina))
                checkbox.grid(row=0, column=0, pady=1)
                entries_HE[i] = tk.Entry(frames[i], width=5)
                entries_DT[i] = tk.Entry(frames[i], width=5)
                entries_TE[i] = tk.Entry(frames[i], width=5)
                entries_HE[i].insert(0, "0")
                entries_DT[i].insert(0, "0") 
                entries_TE[i].insert(0, "0")  
                entries_DT[i].grid(row=1, column=0)
                entries_HE[i].grid(row=0, column=1)
                entries_TE[i].grid(row=1, column=1)
            entry_DF = tk.Entry(frame_scrollable,width=5)
            if len(dias_seleccionados.items())>7:
                entry_comentario = tk.Entry(frame_scrollable)
                entry_comentario.grid(row=index, column=len(dias_seleccionados.items()))
                if checar_aprovacion(empleado_id,periodo[2]):
                        var_aprobar = tk.IntVar(value=1)
                else:
                        var_aprobar = tk.IntVar()
                if fv:
                    # Checkbox "Aprobar"
                    checkbox_aprobar = tk.Checkbutton(frame_scrollable, text="Aprobar", variable=var_aprobar)
                    checkbox_aprobar.grid(row=index, column=len(dias_seleccionados.items())+1)
                    varaprob[empleado_id]=var_aprobar
                    # Botón "Añadir"
                    button_add = tk.Button(frame_scrollable, text="Añadir")
                    button_add.grid(row=index, column=len(dias_seleccionados.items())+2)
                    button_add["command"] = lambda dias=dias_seleccionados,HE_entries=entries_HE,DF=entry_DF,TE_entries=entries_TE,DT_entries=entries_DT, comentario=entry_comentario, var_aprobar=var_aprobar, var2=empleado_id: [desbut(var2),agregar_dato(
                        {dia: var for dia, var in dias.items()},
                        comentario.get(),
                        periodo[2],
                        var_aprobar.get(),
                        var2,{i: HE_entries[i].get() for i in HE_entries},  # Obtiene los valores de HE por cada día
                    DF.get(), 
                    {i: TE_entries[i].get() for i in TE_entries},  # Obtiene los valores de TE por cada día
                    {i: DT_entries[i].get() for i in DT_entries},
                    descanso,Nomina
                    ),desbut(var2)]
                else:
                    # Botón "Añadir"
                    button_add = tk.Button(frame_scrollable, text="Añadir")
                    button_add.grid(row=index, column=len(dias_seleccionados.items())+2)
                    button_add["command"] = lambda dias=dias_seleccionados, HE_entries=entries_HE, DF=entry_DF, TE_entries=entries_TE, DT_entries=entries_DT, comentario=entry_comentario, var_aprobar=var_aprobar, emp_id=empleado_id: agregar_dato(
                    {dia: var.get() for dia, var in dias.items()},
                    comentario.get(),
                    periodo[2],
                    var_aprobar.get(),
                    emp_id, 
                    {i: HE_entries[i].get() for i in HE_entries},  # Obtiene los valores de HE por cada día
                    DF.get(), 
                    {i: TE_entries[i].get() for i in TE_entries},  # Obtiene los valores de TE por cada día
                    {i: DT_entries[i].get() for i in DT_entries},   # Obtiene los valores de DT por cada día
                    descanso,Nomina
                )
            else:
                entry_comentario = tk.Entry(frame_scrollable)
                entry_comentario.grid(row=index, column=7)
                if checar_aprovacion(empleado_id,periodo[2]):
                        var_aprobar = tk.IntVar(value=1)
                else:
                        var_aprobar = tk.IntVar()
                if fv:
                    # Checkbox "Aprobar"
                    checkbox_aprobar = tk.Checkbutton(frame_scrollable, text="Aprobar", variable=var_aprobar)
                    checkbox_aprobar.grid(row=index, column=8)
                    varaprob[empleado_id]=var_aprobar
                    # Botón "Añadir"
                    button_add = tk.Button(frame_scrollable, text="Añadir")
                    button_add.grid(row=index, column=9)
                    button_add["command"] = lambda dias=dias_seleccionados,HE_entries=entries_HE,DF=entry_DF,TE_entries=entries_TE,DT_entries=entries_DT, comentario=entry_comentario, var_aprobar=var_aprobar, var2=empleado_id: [desbut(var2),agregar_dato(
                        {dia: var for dia, var in dias.items()},
                        comentario.get(),
                        periodo[2],
                        var_aprobar.get(),
                        var2,{i: HE_entries[i].get() for i in HE_entries},  # Obtiene los valores de HE por cada día
                    DF.get(), 
                    {i: TE_entries[i].get() for i in TE_entries},  # Obtiene los valores de TE por cada día
                    {i: DT_entries[i].get() for i in DT_entries},
                    descanso,Nomina
                    )]
                else:
                    # Botón "Añadir"
                    button_add = tk.Button(frame_scrollable, text="Añadir")
                    button_add.grid(row=index, column=9)
                    button_add["command"] = lambda dias=dias_seleccionados, HE_entries=entries_HE, DF=entry_DF, TE_entries=entries_TE, DT_entries=entries_DT, comentario=entry_comentario, var_aprobar=var_aprobar, emp_id=empleado_id: agregar_dato(
                    {dia: var for dia, var in dias.items()},
                    comentario.get(),
                    periodo[2],
                    var_aprobar.get(),
                    emp_id, 
                    {i: HE_entries[i].get() for i in HE_entries},  # Obtiene los valores de HE por cada día
                    DF.get(), 
                    {i: TE_entries[i].get() for i in TE_entries},  # Obtiene los valores de TE por cada día
                    {i: DT_entries[i].get() for i in DT_entries},   # Obtiene los valores de DT por cada día
                    descanso,Nomina
                )
            # Otros campos (Comentario, Aprobar, Añadir)
            df=tk.IntVar(value=0) 
            todos.update({empleado_id :{
            'dias': {dia: var for dia, var in dias_seleccionados.items()},
            'comentario': entry_comentario.get(),
            'aprobar': var_aprobar,
            'entries_HE': {i: entries_HE[i].get() for i in entries_HE},
            'entries_DT': {i: entries_DT[i].get() for i in entries_DT},
            'entries_TE': {i: entries_TE[i].get() for i in entries_TE},
            'descanso': descanso,
            'df':df,
            'nomina': Nomina,
            'boton':button_add}
            })
        else:
            index=0
    # Función para añadir todos los empleados
    def añadir_todos():
        global multireg
        multireg=False
        for i,emp_id in enumerate(todos):
            if i==len(todos)-1:
                multireg=True
            agregar_dato(
                todos[emp_id]['dias'],
                todos[emp_id]['comentario'],
                periodo[2],
                todos[emp_id]['aprobar'].get(),
                emp_id,
                todos[emp_id]['entries_HE'],
                todos[emp_id]['df'].get(),
                todos[emp_id]['entries_DT'],
                todos[emp_id]['entries_TE'],
                todos[emp_id]['descanso'],
                todos[emp_id]['nomina']
            )
    btn_añadir_todos = tk.Button(frame_scrollable, text="Añadir para todos",command=lambda:[callback(),añadir_todos()])
    btn_añadir_todos.grid(row=index+1, column=0, columnspan=10)  # Ajusta la posición según sea necesario
# Actualizar el canvas con el tamaño correcto
    frame_scrollable.update_idletasks()
    canvas_scrollable.config(scrollregion=canvas_scrollable.bbox("all"))
    canvas_scrollable.config(width=frame_dinamico.winfo_width(), height=frame_dinamico.winfo_height())
def cerrar_sesion(ventana_principal):
    ventana_principal.destroy()  # Cierra la ventana principal
    mostrar_login()  # Vuelve a mostrar la ventana de inicio de sesión
def cerrar_sesionf(ventana_principal):
    ventana_principal.destroy()  # Cierra la ventana principal
def Agregar_usuario(ventana_principal,):
    ventana_principal.destroy() 
    abrir_ventana_usuarios()  # Vuelve a mostrar la ventana de inicio de sesión
# Función para mostrar la ventana de login nuevamente
def abrir_ventana_usuarios():
    depars=[]
    depars=obtener_ALL_departamentos()
    # Crear la ventana
    ventana_empleados = tk.Tk()
    ventana_empleados.title("Empleados - Registro Usuarios")
    ventana_empleados.state('zoomed')
    usuarios=GET_USER()
    print(usuarios)
    comboboxuser = ttk.Combobox(ventana_empleados, values=usuarios)
    comboboxuser.set(usuarios[0])  # Establecer la opción por defecto
    comboboxuser.pack(pady=5)
    comboboxuser['state'] = "readonly"
    # Botón de cerrar sesión
    button_logout = tk.Button(ventana_empleados, text="Cerrar sesión", command=lambda: cerrar_sesionf(ventana_empleados))
    button_logout.pack(pady=5)
    # Crear una etiqueta para los departamentos
    label_departamentos = tk.Label(ventana_empleados, text="Selecciona los departamentos:")
    label_departamentos.pack(pady=10)

    # Agregar campo de búsqueda y botón de búsqueda
    search_frame = tk.Frame(ventana_empleados)
    search_frame.pack(pady=5)
    search_entry = tk.Entry(search_frame, width=30)
    search_entry.pack(side="left", padx=5)
    
    # Crear un contenedor con desplazamiento para los checkboxes
    frame_con_scroll = tk.Frame(ventana_empleados)
    frame_con_scroll.pack(pady=5, fill='both', expand=True)

    # Crear un Canvas para contener el Frame de departamentos
    canvas = tk.Canvas(frame_con_scroll)
    scrollbar = ttk.Scrollbar(frame_con_scroll, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    # Configurar el frame interno en el canvas
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Colocar el canvas y el scrollbar
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Lista para almacenar los estados de las checkboxes
    checkbox_vars = []
    checkboxes = []
    vars=[]
     # Función que se ejecuta al cambiar el estado de cada checkbox
    def actualizar_valor(departamento, var):
        for i, (dep, _) in enumerate(checkbox_vars):
            if dep == departamento:
                print(var.get())
                checkbox_vars[i] = (dep, var)
                print(checkbox_vars[i])
                break

    # Crear las checkboxes y vincular la función de actualización
    def crear():
        search_text = search_entry.get().lower()
        for i, (departamento, _) in enumerate(checkbox_vars):
            if search_text in departamento.lower():
                checkboxes[i].pack_forget()
            else:
                checkboxes[i].pack_forget()
        checkboxes.clear()
        checkbox_vars.clear()
        for i,departamento in enumerate(depars):
            vars.append(tk.IntVar())
            checkbox_vars.append((departamento, vars[i]))  # Añadir a la lista
        checkdep(comboboxuser.get(),checkbox_vars,vars)
        for i,departamento in enumerate(depars):
            checkbox = tk.Checkbutton(
                scrollable_frame, text=departamento, variable=vars[i],
                command=lambda d=departamento, v=vars[i]: actualizar_valor(d, v)
            )
            checkbox.pack(anchor='w')
            checkboxes.append(checkbox)
    crear()
    # Función para buscar y filtrar departamentos
    def buscar_departamento():
        search_text = search_entry.get().lower()
        for i, (departamento, _) in enumerate(checkbox_vars):
            if search_text in departamento.lower():
                checkboxes[i].pack(anchor='w')
            else:
                checkboxes[i].pack_forget()

    # Botón de búsqueda
    search_button = tk.Button(search_frame, text="Buscar", command=buscar_departamento)
    search_button.pack(side="left", padx=5)

    # Función para desplazarse con la rueda del ratón
    def on_mouse_wheel(event):
        if event.num == 4:  # Para sistemas que usan Button-4 y Button-5
            canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            canvas.yview_scroll(1, "units")
        else:  # Para Windows y otros sistemas con MouseWheel
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    # Vincular la rueda del ratón con el desplazamiento
    canvas.bind_all("<MouseWheel>", on_mouse_wheel)        # Windows
    canvas.bind_all("<Button-4>", on_mouse_wheel)          # Otros sistemas
    canvas.bind_all("<Button-5>", on_mouse_wheel)          # Otros sistemas
    # Botón para imprimir los departamentos seleccionados
    def mostrar_seleccion():
        print([(dep,var.get()) for dep, var in checkbox_vars])
        seleccionados = [dep for dep, var in checkbox_vars if var.get()]
        print("Departamentos seleccionados:", seleccionados)
        UPDATE_USER(comboboxuser.get(),seleccionados)
    comboboxuser.bind("<<ComboboxSelected>>", lambda event: crear())
    button_seleccion = tk.Button(ventana_empleados, text="Mostrar Selección", command=mostrar_seleccion)
    button_seleccion.pack(pady=10)
    ventana_empleados.grid_rowconfigure(1, weight=1)
    ventana_empleados.grid_columnconfigure(0, weight=1)
    ventana_empleados.mainloop()
def mostrar_login():
    global root, entry_username, entry_password
    root = tk.Tk()
    root.title("Inicio de Sesión")
    root.state('zoomed') 
    # Etiqueta y campo para el nombre de usuario
    label_username = tk.Label(root, text="Nombre de Usuario:")
    label_username.pack(pady=5)
    entry_username = tk.Entry(root)
    entry_username.pack(pady=5)
    # Etiqueta y campo para la contraseña
    label_password = tk.Label(root, text="Contraseña:")
    label_password.pack(pady=5)
    entry_password = tk.Entry(root, show='*')
    entry_password.pack(pady=5)
    # Botón para iniciar sesión
    button_login = tk.Button(root, text="Iniciar Sesión", command=iniciar_sesion)
    button_login.pack(pady=20)
    root.mainloop()
# Mostrar la ventana de inicio de sesión al inicio del programa
excel()
mostrar_login()